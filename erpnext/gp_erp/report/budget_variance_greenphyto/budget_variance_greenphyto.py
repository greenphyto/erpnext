# Copyright (c) 2015, Frappe Technologies Pvt. Ltd. and Contributors
# License: GNU General Public License v3. See license.txt


import frappe
import calendar
import re
import json
from urllib.parse import quote
from io import BytesIO
try:
	import openpyxl
except Exception:
	openpyxl = None
from frappe import _
from frappe.utils import flt, now_datetime, get_datetime, now, cint, get_first_day, get_last_day, getdate

from erpnext.accounts.report.financial_statements import (
	get_columns,
	get_data,
	get_filtered_list_for_consolidated_report,
	get_period_list,
)
from erpnext.accounts.utils import remove_account_number
from erpnext.accounts.report.utils import convert_wrap_report_data


def execute(filters=None):
	filters = control_filters(filters)
	period_list = get_period_list(
		filters.from_fiscal_year,
		filters.to_fiscal_year,
		filters.period_start_date,
		filters.period_end_date,
		filters.filter_based_on,
		filters.periodicity,
		company=filters.company,
		month=filters.month,
		to_month=filters.to_month,
		# ytd=1
	)

	accounts_list = []
	if cint(filters.hide_zero_balance):
		accounts_list = get_budget_account(filters.get("cost_center"), filters.get("company"))

	income = get_data(
		filters.company,
		"Income",
		"Credit",
		period_list,
		filters=filters,
		accumulated_values=filters.accumulated_values,
		ignore_closing_entries=True,
		ignore_accumulated_values_for_fy=True,
		filter_zero_value=0,
		accounts_to_show=accounts_list
	)

	expense = get_data(
		filters.company,
		"Expense",
		"Debit",
		period_list,
		filters=filters,
		accumulated_values=filters.accumulated_values,
		ignore_closing_entries=True,
		ignore_accumulated_values_for_fy=True,
		filter_zero_value=0,
		accounts_to_show=accounts_list
	)

	# Fetch budget data and calculate YTD (Year-to-Date) BEFORE extending to data
	budget_map = {}
	# ALWAYS fetch monthly budget data (even for Yearly periodicity)
	# This allows Budget YTD to calculate from raw monthly values
	budget_map = get_budget_data(filters, ytd=False)

	# Build total company map (same report scope but without cost center filter)
	# total_company_map = get_total_company_map(filters, period_list, accounts_list)
	
	# Add budget to income rows (includes budget columns and summary columns)
	if income:
		add_budget_to_rows(income, budget_map, period_list, filters)
		add_summary_columns(income, period_list)
		# add_total_company_column(income, total_company_map)
	
	# Add budget to expense rows (includes budget columns and summary columns)
	if expense:
		add_budget_to_rows(expense, budget_map, period_list, filters)
		add_summary_columns(expense, period_list)
		# add_total_company_column(expense, total_company_map)
	
	# Calculate net profit/loss AFTER budget is added so we can include budget columns
	net_profit_loss = get_net_profit_loss(
		income, expense, period_list, filters.company, filters.presentation_currency
	)

	
	data = []
	data.extend(income or [])
	data.extend(expense or [])
	if net_profit_loss:
		data.append(net_profit_loss)

	new_data = []
	if filters.show_number_group:
		new_data = data
	else:
		for d in data:
			if d.get("is_group"):
				d['account_name'] = remove_account_number(d['account_name'])
				if frappe.flags.in_export:
					d['account'] = d['account_name']
					if not d.get('parent_account'):
						for key, val in d.items():
							if key not in ['account', 'account_name']:
								d[key] = None

			new_data.append(d)
	# new_data.pop()
	filters.show_budget_amount = 1
	# filters.ytd_column = 1
	columns = get_report_column(filters, period_list)

	chart = None #get_chart_data(filters, columns, income, expense, net_profit_loss)

	report_summary = None

	return columns, new_data, None, chart, report_summary

def control_filters(filters):
	start_date = getdate("{}-{}-01".format(filters.year, filters.month))
	end_date = getdate("{}-{}-01".format(filters.year, filters.to_month))
	filters = frappe._dict(filters)
	filters.from_fiscal_year = filters.year
	filters.to_fiscal_year = filters.year
	filters.period_start_date = get_first_day(start_date)
	filters.period_end_date = get_last_day(end_date)
	return filters

def get_budget_account(cost_center=None, company=None):
	"""
	Get unique accounts from Budget Account that match the cost center and company filters.
	
	Args:
		cost_center: List of cost centers or None
		company: Company name or None
	
	Returns:
		List of unique account names
	"""
	conditions = ""
	filters = frappe._dict({})
	
	# Handle cost_center as a list
	if cost_center:
		if isinstance(cost_center, str):
			cost_center = json.loads(cost_center)
		conditions += " AND b.cost_center in %(cost_center)s "
		filters.cost_center = cost_center

	if company:
		conditions += " AND b.company = %(company)s "
		filters.company = company
	
	# Query with JOIN to Budget parent table
	query = f"""
		SELECT DISTINCT ba.account 
		FROM `tabBudget Account` ba
		INNER JOIN `tabBudget` b ON ba.parent = b.name
		AND b.docstatus = 1
		{conditions}
		ORDER BY ba.account
	"""
	
	return frappe.db.sql(query, filters, pluck="account")

def get_report_column(filters, period_list):
	temp_columns = get_columns(
		filters.periodicity, period_list, filters.accumulated_values, filters.company, cost_center_all_show=filters.show_all_cost_centers, filters=filters
	)

	# skip total column bcs use total actual
	columns = []
	for col in temp_columns:
		fieldname = col.get("fieldname", "")
		if fieldname == "total":
			continue
		else:
			print(fieldname) 
			if filters.periodicity == "Yearly":
				if fieldname in ['account', 'acc_code', 'currency']:
					columns.append(col)
			else:
				columns.append(col)


	# Total Actual column
	# Convert full month names to abbreviated names (e.g., 'January' -> 'Jan')
	def month_abbr(month_name):
		if not month_name:
			return ""
		try:
			month_num = list(calendar.month_name).index(month_name)
			if month_num == 0:
				return month_name[:3]
			return calendar.month_abbr[month_num]
		except Exception:
			return month_name[:3]

	from_month_abbr = month_abbr(getattr(filters, 'month', None))
	to_month_abbr = month_abbr(getattr(filters, 'to_month', None))
	period_word = f"{from_month_abbr} - {to_month_abbr} {getattr(filters, 'year', '')}"
	columns.append({
		"fieldname": "total_actual",
		"label": _(period_word+ " (Actual)"),
		"fieldtype": "Currency",
		"options": "currency",
		"width": 200,
		"align": "right",
	})

	# Total Company column (without cost center filter)
	# columns.append({
	# 	"fieldname": "total_company",
	# 	"label": _("Total Company"),
	# 	"fieldtype": "Currency",
	# 	"options": "currency",
	# 	"width": 150,
	# 	"align": "right",
	# })
	
	# Budget YTD column
	columns.append({
		"fieldname": "budget_ytd",
		"label": _(period_word+ " (Budget)"),
		"fieldtype": "Currency",
		"options": "currency",
		"width": 200,
		"align": "right",
	})
	
	# Variance $ column
	columns.append({
		"fieldname": "variance_amount",
		"label": _(period_word+ " (Variance $)"),
		"fieldtype": "Currency",
		"options": "currency",
		"width": 200,
		"align": "right",
	})
	
	# Variance % column
	columns.append({
		"fieldname": "variance_percent",
		"label": _(period_word+ " (Variance %)"),
		"fieldtype": "Percent",
		"width": 210,
		"align": "right",
	})

	return columns

def get_report_summary(
	period_list, periodicity, income, expense, net_profit_loss, currency, filters, consolidated=False
):
	net_income, net_expense, net_profit = 0.0, 0.0, 0.0

	# from consolidated financial statement
	if filters.get("accumulated_in_group_company"):
		period_list = get_filtered_list_for_consolidated_report(filters, period_list)

	for period in period_list:
		key = period if consolidated else period.key
		if income:
			net_income += income[-2].get(key)
		if expense:
			net_expense += expense[-2].get(key)
		if net_profit_loss:
			net_profit += net_profit_loss.get(key)

	if len(period_list) == 1 and periodicity == "Yearly":
		profit_label = _("Profit This Year")
		income_label = _("Total Income This Year")
		expense_label = _("Total Expense This Year")
	else:
		profit_label = _("Net Profit")
		income_label = _("Total Income")
		expense_label = _("Total Expense")

	return [
		{"value": net_income, "label": income_label, "datatype": "Currency", "currency": currency},
		{"type": "separator", "value": "-"},
		{"value": net_expense, "label": expense_label, "datatype": "Currency", "currency": currency},
		{"type": "separator", "value": "=", "color": "blue"},
		{
			"value": net_profit,
			"indicator": "Green" if net_profit > 0 else "Red",
			"label": profit_label,
			"datatype": "Currency",
			"currency": currency,
		},
	]


def get_net_profit_loss(income, expense, period_list, company, currency=None, consolidated=False):
	total = 0
	net_profit_loss = {
		"account_name": "'" + _("Profit for the year") + "'",
		"account": "'" + _("Profit for the year") + "'",
		"warn_if_negative": True,
		"profit_data":1,
		"currency": currency or frappe.get_cached_value("Company", company, "default_currency"),
	}

	has_value = False

	for period in period_list:
		key = period if consolidated else period.key
		total_income = flt(income[-2][key], 3) if income else 0
		total_expense = flt(expense[-2][key], 3) if expense else 0

		net_profit_loss[key] = total_income - total_expense
		
		# Also calculate budget for net profit/loss (income budget - expense budget)
		budget_key = (period if consolidated else period.key) + "_budget" if not consolidated else None
		if budget_key:
			income_budget = flt(income[-2].get(budget_key, 0), 3) if income else 0
			expense_budget = flt(expense[-2].get(budget_key, 0), 3) if expense else 0
			net_profit_loss[budget_key] = income_budget - expense_budget

		if net_profit_loss[key]:
			has_value = True

		total += flt(net_profit_loss[key])
		net_profit_loss["total"] = total

	# Add summary columns to net profit/loss row
	if net_profit_loss and income and expense:
		# Calculate total_actual, budget_ytd, variance for net profit/loss
		income_total = income[-2] if income else {}
		expense_total = expense[-2] if expense else {}
		
		# Total Actual = Income Total Actual - Expense Total Actual
		net_profit_loss["total_actual"] = flt(income_total.get("total_actual", 0)) - flt(expense_total.get("total_actual", 0))

		# Total Company = Income Total Company - Expense Total Company
		net_profit_loss["total_company"] = flt(income_total.get("total_company", 0)) - flt(expense_total.get("total_company", 0))
		
		# Budget YTD = Income Budget YTD - Expense Budget YTD (only periods up to current date)
		net_profit_loss["budget_ytd"] = flt(income_total.get("budget_ytd", 0)) - flt(expense_total.get("budget_ytd", 0))
		
		# Variance Amount = Total Actual - Budget YTD
		net_profit_loss["variance_amount"] = net_profit_loss["total_actual"] - net_profit_loss["budget_ytd"]
		
		# Variance % = (Variance Amount / |Budget YTD|) × 100
		# Use absolute value so that the sign is not reversed when the budget is negative (loss)
		budget_ytd = net_profit_loss["budget_ytd"]
		if budget_ytd != 0:
			net_profit_loss["variance_percent"] = flt((net_profit_loss["variance_amount"] / flt(budget_ytd)) * 100, 2)
		else:
			net_profit_loss["variance_percent"] = 0

	if has_value:
		return net_profit_loss


def get_chart_data(filters, columns, income, expense, net_profit_loss):
	# Filter out budget columns from chart data
	period_columns = [d for d in columns[2:] if not d.get("fieldname", "").endswith("_budget")]
	
	labels = [d.get("label") for d in period_columns]

	income_data, expense_data, net_profit = [], [], []

	for p in period_columns:
		if income:
			income_data.append(income[-2].get(p.get("fieldname")))
		if expense:
			expense_data.append(expense[-2].get(p.get("fieldname")))
		if net_profit_loss:
			net_profit.append(net_profit_loss.get(p.get("fieldname")))

	datasets = []
	if income_data:
		datasets.append({"name": _("Income"), "values": income_data})
	if expense_data:
		datasets.append({"name": _("Expense"), "values": expense_data})
	if net_profit:
		datasets.append({"name": _("Net Profit/Loss"), "values": net_profit})

	chart = {"data": {"labels": labels, "datasets": datasets}}

	if not cint(filters.accumulated_values):
		chart["type"] = "bar"
	else:
		chart["type"] = "line"

	chart["fieldtype"] = "Currency"

	return chart

from frappe.desk.query_report import add_title_report, get_filters_data, build_xlsx_data
from openpyxl.utils import get_column_letter
def get_export_cost_center(filters):
	"""
	Generate grouped Profit & Loss data per non-group Cost Center.

	Returns a dict keyed by Cost Center name with:
	- label: Cost Center display name
	- columns: report columns for this run
	- data: report rows for this Cost Center
	- summary: report summary values
	"""
	base_filters = frappe._dict(filters or {})

	# Collect all non-group cost centers (scoped to company if provided)
	cc_filters = {"is_group": 0}
	if base_filters.get("company"):
		cc_filters["company"] = base_filters.company

	cost_centers = frappe.get_all(
		"Cost Center",
		filters=cc_filters,
		fields=["name", "cost_center_name"],
		order_by="name asc",
	)

	group_data = {}

	# add detail
	export_date = now()
	date_str = " "+get_datetime(export_date).strftime("%-d %B %y %H:%M:%S")
	report_name = "Profit and Loss Statement"
	title_report = add_title_report(report_name) 
	filter_report = get_filters_data(filters)

	for idx, cc in enumerate(cost_centers):
		# Prepare per-CC filters without mutating caller filters
		per_cc_filters = frappe._dict(base_filters.copy())
		# empty cost center at first time
		if idx == 0:
			per_cc_filters["cost_center"] = []
		else:
			per_cc_filters["cost_center"] = [cc.name]

		columns, data, _, _, report_summary = execute(per_cc_filters)
		temp = frappe._dict({
			"columns":columns,
			"result":data
		})
		xlsx_data, column_widths = build_xlsx_data(temp, [] , 1, ignore_visible_idx=1)
		xlsx_data = title_report + filter_report + [["Export date", date_str]] + xlsx_data

		group_data[cc.name] = {
			"label": cc.cost_center_name or cc.name,
			"columns": columns,
			"data": xlsx_data,
			"summary": report_summary,
			"column_widths":column_widths
		}

	return group_data


def _sanitize_sheet_name(name):
	name = (name or "Sheet").strip()
	# Excel invalid chars: : \ / ? * [ ]
	name = re.sub(r"[:\\/\?\*\[\]]", " ", name)
	if len(name) > 31:
		name = name[:31]
	return name or "Sheet"


@frappe.whitelist()
def export_with_cost_centers(filters=None):
	"""
	Build an XLSX where each Cost Center is a sheet containing its P&L data.

	Returns: { file_url: <saved file url> }
	"""
	if isinstance(filters, str):
		try:
			filters = frappe.parse_json(filters)
		except Exception:
			filters = {}

	filters['show_all_cost_centers'] = 0
	group_data  = get_export_cost_center(filters)

	if not openpyxl:
		frappe.throw("openpyxl is required to export XLSX")

	from frappe.utils.xlsxutils import ILLEGAL_CHARACTERS_RE, handle_html

	wb = openpyxl.Workbook(write_only=True)

	used_names = set()
	first_columns = None

	for cc_name, payload in (group_data or {}).items():
		label = payload.get("label") or cc_name
		sheet_name = _sanitize_sheet_name(label)

		# Ensure unique sheet names
		base = sheet_name
		idx = 1
		while sheet_name in used_names:
			suffix = f" {idx}"
			sheet_name = _sanitize_sheet_name((base[: (31 - len(suffix))]).rstrip() + suffix)
			idx += 1
		used_names.add(sheet_name)

		ws = wb.create_sheet(title=sheet_name)

		columns = payload.get("columns") or []
		if first_columns is None:
			first_columns = columns
		data_rows = payload.get("data") or []

		headers = [c.get("label") for c in columns]
		fields = [c.get("fieldname") for c in columns]
		ws.append(headers)

		# Set column width if provided
		column_widths = payload.get("column_widths")
		for i, column_width in enumerate(column_widths):
			if column_width:
				ws.column_dimensions[get_column_letter(i + 1)].width = column_width

		for row in data_rows:
			out = []
			if type(row) != list:
				for f in fields:
					val = (row or {}).get(f)
					if isinstance(val, str):
						try:
							val = handle_html(val)
							if isinstance(val, str):
								val = re.sub(ILLEGAL_CHARACTERS_RE, "", val)
						except Exception:
							pass
					out.append(val)
			else:
				out += row

			ws.append(out)

	bio = BytesIO()
	wb.save(bio)
	bio.seek(0)

	# Integrate with add_formulas to add formulas on all sheets
	try:
		from erpnext.accounts.report.balance_sheet_v2.balance_sheet_v2 import add_formulas
		# Build column_widths list from the first sheet's columns if available
		column_widths = None
		if first_columns:
			column_widths = []
			for col in first_columns:
				w = col.get("width")
				# Convert typical pixel widths to character widths approximation if needed
				if isinstance(w, (int, float)):
					# heuristic: 1 char ~ 8 px; clamp sensible range
					width_chars = max(6, min(80, int(float(w) / 8)))
				else:
					# fall back to label length approx
					label = col.get("label") or ""
					width_chars = max(8, min(50, len(str(label)) + 5))
				column_widths.append(width_chars)

		return add_formulas("Profit and Loss Statement", bio, column_widths=column_widths)

	except Exception:
		# Fallback: plain export if add_formulas unavailable
		now = now_datetime()
		date_str_title = now.strftime("%y%m%d_%H%M%S")
		frappe.response["filename"] = f"Profit and Loss by Cost Center_{date_str_title}.xlsx"
		frappe.response["filecontent"] = bio.getvalue()
		frappe.response["type"] = "binary"


def get_budget_data(filters, ytd=False):
	"""
	Fetch budget data for accounts based on fiscal year, company, and cost center.
	Budget is fetched directly from monthly fields (january, february, etc.) in Budget Account child table.
	
	Returns:
		- If periodicity is "Monthly": dict mapping account -> month_number -> budget_amount
		- If periodicity is "Yearly": dict mapping account -> fiscal_year -> total_budget_amount
	"""

	periodicity = filters.get("periodicity")
	
	budget_against = frappe.scrub(filters.get("budget_against", "Cost Center"))
	cost_centers = filters.get("cost_center") or []
	
	# Build condition for cost centers
	cond = ""
	if cost_centers:
		cond += """ and b.{budget_against} in (%s)""".format(budget_against=budget_against) % ", ".join(
			["%s"] * len(cost_centers)
		)
	
	# Fetch budget details directly from Budget Account with monthly fields
	budget_details = frappe.db.sql(
		"""
			select
				ba.account,
				sum(coalesce(ba.january, 0)) as january,
				sum(coalesce(ba.february, 0)) as february,
				sum(coalesce(ba.march, 0)) as march,
				sum(coalesce(ba.april, 0)) as april,
				sum(coalesce(ba.may, 0)) as may,
				sum(coalesce(ba.june, 0)) as june,
				sum(coalesce(ba.july, 0)) as july,
				sum(coalesce(ba.august, 0)) as august,
				sum(coalesce(ba.september, 0)) as september,
				sum(coalesce(ba.october, 0)) as october,
				sum(coalesce(ba.november, 0)) as november,
				sum(coalesce(ba.december, 0)) as december
			from
				`tabBudget` b
				inner join `tabBudget Account` ba on b.name = ba.parent
			where
				b.docstatus = 1
				and b.company = %s
				and b.fiscal_year = %s
				and b.budget_against = %s
				{cond}
			group by
				ba.account
			order by
				ba.account
		""".format(
			cond=cond,
		),
		tuple(
			[
				filters.get("company"),
				filters.get("from_fiscal_year"),
				filters.get("budget_against", "Cost Center"),
			] + cost_centers
		),
		as_dict=True,
	)
	
	# Month mapping: month_number -> field_name
	month_fields = {
		1: "january",
		2: "february",
		3: "march",
		4: "april",
		5: "may",
		6: "june",
		7: "july",
		8: "august",
		9: "september",
		10: "october",
		11: "november",
		12: "december"
	}
	
	budget_map = {}
	current_month = None
	if ytd:
		current_month = get_datetime().month
		
	for bd in budget_details:
		account = bd.account
		if account not in budget_map:
			budget_map[account] = {}
		
		for month_num in range(1, 13):
			# if ytd and month_num > current_month:
			# 	break

			month_field = month_fields[month_num]
			monthly_budget = flt(bd.get(month_field, 0))
			
			key = month_num
			
			if key not in budget_map[account]:
				budget_map[account][key] = 0
			budget_map[account][key] += monthly_budget
	
	return budget_map


def add_budget_to_rows(rows, budget_map, period_list, filters):
	"""Add budget amounts to all rows including totals
	
	Args:
		rows: List of account rows (income or expense)
		budget_map: Dict mapping account -> period -> budget_amount
		period_list: List of periods to add budget for
		current_date: Current date for YTD limit
		filters: Report filters
	"""
	from frappe.utils import getdate, add_months, flt
	
	# First pass: Add budget to leaf accounts (non-group, non-total)
	for row in rows:
		# Skip total rows (they'll be calculated later)
		account_name = row.get("account_name", "").lower()
		is_total_row = "total" in account_name and ("income" in account_name or "expense" in account_name)
		
		if is_total_row:
			continue
		
		# Skip group accounts in first pass
		if row.get("is_group"):
			continue
			
		account = row.get("account_origin") or row.get("account")
		if not account:
			continue
		
		# Add budget for each period
		for period in period_list:
			period_from_date = getdate(period.from_date) if period.from_date else None
			period_to_date = getdate(period.to_date) if period.to_date else None

			# skip if outside filter date
			if period_to_date and filters.get("period_end_date") and period_to_date > getdate(filters.get("period_end_date")):
				continue
			if period_from_date and filters.get("period_start_date") and period_from_date < getdate(filters.get("period_start_date")):
				continue

			budget_key = period.key + "_budget"
			budget_value = 0
			
			if filters.periodicity == "Yearly":
				# For yearly budget, sum all 12 months from budget_map
				# budget_map is now always monthly: account -> month_number -> budget_amount
				if account in budget_map:
					for month_num in range(1, 13):
						budget_value += flt(budget_map.get(account, {}).get(month_num, 0))
			else:
				# Monthly budget
				if account in budget_map and period_to_date:
					if filters.accumulated_values:
						# Accumulated mode: Calculate YTD
						fiscal_year_start = getdate(period.year_start_date) if period.year_start_date else None
						
						if fiscal_year_start:
							ytd_budget = 0
							current_month_date = fiscal_year_start
							
							# Iterate month by month from fiscal year start to period end
							month_num = current_month_date.month
							ytd_budget += budget_map.get(account, {}).get(month_num, 0)
							
							# Move to next month
							current_month_date = add_months(current_month_date, 1)
							
							# Safety check
							if current_month_date.year > period_to_date.year or \
								(current_month_date.year == period_to_date.year and current_month_date.month > period_to_date.month):
								break
							
							budget_value = ytd_budget
					else:
						# Non-accumulated mode: Show budget for this month only
						month_num = period_to_date.month
						budget_value = budget_map.get(account, {}).get(month_num, 0)

			row[budget_key] = budget_value
	
	# Second pass: Calculate budget for group accounts (bottom-up)
	total_root_group = {}
	for row in reversed(rows):
		account_name = row.get("account_name", "").lower()
		is_total_row = "total" in account_name and ("income" in account_name or "expense" in account_name)
		
		if is_total_row:
			total_root_group = row
			continue

		if row.get("is_group"):
			idx = rows.index(row)
		
			account = row.get("account_origin") or row.get("account")
			if not account:
				continue
			
			# Initialize budget totals for this group
			for period in period_list:
				budget_key = period.key + "_budget"
				row[budget_key] = 0

				if idx == 0 and total_root_group:
					# If this is the top-level group, also initialize totals in total_root_group
					total_root_group[budget_key] = 0
			
			# Sum budget from child accounts
			for child_row in rows:
				child_account_name = child_row.get("account_name", "").lower()
				child_is_total = "total" in child_account_name and ("income" in child_account_name or "expense" in child_account_name)
				
				if child_is_total:
					continue
				
				child_parent = child_row.get("parent_account")
				if child_parent == account:
					for period in period_list:
						budget_key = period.key + "_budget"
						child_budget = flt(child_row.get(budget_key, 0))
						row[budget_key] = flt(row.get(budget_key, 0)) + child_budget

						if idx == 0 and total_root_group:
							total_root_group[budget_key] = flt(total_root_group.get(budget_key, 0)) + child_budget

def add_summary_columns(rows, period_list):
	"""Add summary columns: Total Actual, Budget YTD, Variance $, Variance %
	
	Args:
		rows: List of account rows (income or expense)
		budget_map: Raw budget data (account -> month -> amount)
		period_list: List of periods 
		current_date: Current date for YTD limit
		filters: Report filters
	"""
	from frappe.utils import flt, getdate, cint, add_months
	
	for row in rows:
		# Skip profit/loss row
		if row.get("profit_data"):
			continue
		
		# Calculate Total Actual - sum of all period values
		total_actual = 0
		for period in period_list:
			period_value = flt(row.get(period.key, 0))
			total_actual += period_value
		
		row["total_actual"] = total_actual
		
		# Calculate Budget YTD - ALWAYS from raw monthly budget data up to current month
		# This ensures correct calculation regardless of periodicity (Monthly/Yearly)
		budget_ytd = 0
		account = row.get("account_origin") or row.get("account")
		
		for period in period_list:
			key_budget = period.key + "_budget"
			budget_ytd += flt(row.get(key_budget, 0))
		
		row["budget_ytd"] = budget_ytd
		
		# Calculate Variance $ = Total Actual - Budget YTD
		variance_amount = total_actual - budget_ytd
		row["variance_amount"] = variance_amount
		
		# Calculate Variance % = (Variance $ / Budget YTD) × 100
		if budget_ytd != 0:
			variance_percent = (variance_amount / budget_ytd) * 100
		else:
			variance_percent = 0
		
		row["variance_percent"] = flt(variance_percent, 2)

def get_total_company_map(filters, period_list, accounts_list=None):
	"""Get total actual per account without cost center filter.

	Returns:
		dict: account -> total amount across selected periods
	"""
	accounts_list = accounts_list or []
	total_company_map = {}

	# If report is already not filtered by cost center, total company equals current total actual.
	if not filters.get("cost_center"):
		return total_company_map

	filters_no_cost_center = frappe._dict(filters.copy())
	filters_no_cost_center.cost_center = []

	for root_type, balance_must_be in (("Income", "Credit"), ("Expense", "Debit")):
		rows = get_data(
			filters_no_cost_center.company,
			root_type,
			balance_must_be,
			period_list,
			filters=filters_no_cost_center,
			accumulated_values=filters_no_cost_center.accumulated_values,
			ignore_closing_entries=True,
			ignore_accumulated_values_for_fy=True,
			filter_zero_value=0,
			accounts_to_show=accounts_list,
		)

		for row in rows or []:
			account = row.get("account_origin") or row.get("account")
			if not account:
				continue

			total_actual = 0
			for period in period_list:
				total_actual += flt(row.get(period.key, 0))

			total_company_map[account] = total_actual

	return total_company_map


def add_total_company_column(rows, total_company_map=None):
	# add total amount account balance under company on new column
	# column = total_company
	# this total is account's balance without cost center filter,
	# so it shows total balance of the account for the company
	total_company_map = total_company_map or {}

	for row in rows or []:
		account = row.get("account_origin") or row.get("account")
		if account and total_company_map:
			row["total_company"] = flt(total_company_map.get(account, 0))
		else:
			# Fallback: when no cost center filter is applied, this equals current total actual
			row["total_company"] = flt(row.get("total_actual", 0))

@frappe.whitelist()
def get_export_with_cost_centers_url(filters=None):
	"""
	Helper to generate a URL for binary export so that the client
	can use frappe.call() first, then window.open() the returned URL.
	"""
	if isinstance(filters, str):
		try:
			filters = frappe.parse_json(filters)
		except Exception:
			filters = {}

	payload = quote(json.dumps(filters or {}))
	url = (
		"/api/method/erpnext.accounts.report.profit_and_loss_statement.profit_and_loss_statement.export_with_cost_centers"
		f"?filters={payload}&report_name=Budget%20Variance%20Greenphyto&formula=0"
	)
	return {"url": url}