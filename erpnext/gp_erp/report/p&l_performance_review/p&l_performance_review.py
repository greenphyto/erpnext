# Copyright (c) 2026, Frappe Technologies Pvt. Ltd. and contributors
# For license information, please see license.txt

import calendar

import frappe
from frappe import _
from frappe.utils import (
	add_years,
	cint,
	flt,
	get_first_day,
	get_last_day,
	getdate,
	nowdate,
)

from erpnext.accounts.report.financial_statements import get_data, get_period_list
from erpnext.accounts.utils import get_fiscal_year, remove_account_number
from erpnext.gp_erp.report.budget_variance_greenphyto.budget_variance_greenphyto import (
	add_budget_to_rows,
	add_summary_columns,
	get_budget_account,
	get_budget_data,
)

MONTH_NAMES = list(calendar.month_name)

ACCOUNT_GROUPS = {
	"direct_income": "Direct Income",
	"indirect_income": "Indirect Income",
	"cogs": "Cost of Goods Sold",
	"cos": "Cost of Sales",
	"opex": "Expenses-Operating",
	"depreciation": "Depreciation",
	"finance": "Finance Expenses",
	"tax": "Tax Expenses- GPL",
}

OPEX_SUB_GROUPS = [
	("600001", "Staff Cost"),
	("600002", "Selling & Marketing"),
	("600003", "Audit & Consultation Fees"),
	("600004", "Repairs and Maintenance"),
	("600005", "Operating Expense"),
	("600006", "Subscription Fees"),
	("600007", "Realised FX (Gain) / Loss"),
]


def execute(filters=None):
	filters = control_filters(filters)

	period_list = get_period_list(
		filters.from_fiscal_year,
		filters.to_fiscal_year,
		filters.period_start_date,
		filters.period_end_date,
		filters.get("filter_based_on"),
		filters.periodicity,
		company=filters.company,
	)

	to_m = list(MONTH_NAMES).index(filters.to_month)
	period_list = [p for p in period_list if p.to_date.month <= to_m]
	display_period_list = period_list[-1:]

	accounts_list = []
	if cint(filters.hide_zero_balance):
		accounts_list = get_budget_account(filters.get("cost_center"), filters.get("company"))

	income, expense = fetch_pl(filters, period_list, accounts_list)

	budget_map = get_budget_data(filters, ytd=False)
	if income:
		add_budget_to_rows(income, budget_map, period_list, filters)
		add_summary_columns(income, period_list)
	if expense:
		add_budget_to_rows(expense, budget_map, period_list, filters)
		add_summary_columns(expense, period_list)

	_month_key = display_period_list[-1].key
	for _rows in (income, expense):
		fill_monthly_variance(_rows, _month_key)

	merge_prior_year(income, expense, filters, period_list, accounts_list, _month_key)

	for row in (income or []) + (expense or []):
		if row:
			fill_prior_month_var([row], _month_key)

	new_data = restructure_for_display(income, expense, filters, period_list, _month_key)
	columns = get_report_columns(filters, display_period_list)
	return columns, new_data


def control_filters(filters):
	filters = frappe._dict(filters or {})
	if not filters.get("year"):
		frappe.throw(_("Year is required"))

	# Always monthly system — no YTD / Accumulated toggle
	filters.view_mode = "Monthly"
	filters.periodicity = "Monthly"
	filters.accumulated_values = 0
	filters.display_accumulated = 0

	filters.month = "January"
	current = getdate(nowdate())
	try:
		year_num = cint(str(filters.year)[:4])
	except Exception:
		year_num = current.year

	if filters.get("to_month") in MONTH_NAMES:
		pass
	elif year_num < current.year:
		filters.to_month = "December"
	else:
		filters.to_month = MONTH_NAMES[current.month]

	start_date = getdate("{}-01-01".format(year_num))
	end_date = getdate("{}-{}-01".format(year_num, list(MONTH_NAMES).index(filters.to_month)))
	filters.from_fiscal_year = filters.year
	filters.to_fiscal_year = filters.year
	filters.period_start_date = get_first_day(start_date)
	filters.period_end_date = get_last_day(end_date)
	filters.filter_based_on = "Fiscal Year"
	filters.budget_against = filters.get("budget_against") or "Cost Center"
	return filters


def fetch_pl(filters, period_list, accounts_list=None):
	accounts_list = accounts_list or []
	income = get_data(
		filters.company,
		"Income",
		"Credit",
		period_list,
		filters=filters,
		accumulated_values=filters.accumulated_values,
		ignore_closing_entries=True,
		ignore_accumulated_values_for_fy=True,
		filter_zero_value=0,
		accounts_to_show=accounts_list,
	)
	expense = get_data(
		filters.company,
		"Expense",
		"Debit",
		period_list,
		filters=filters,
		accumulated_values=filters.accumulated_values,
		ignore_closing_entries=True,
		ignore_accumulated_values_for_fy=True,
		filter_zero_value=0,
		accounts_to_show=accounts_list,
	)
	return income, expense


def fill_monthly_variance(rows, month_key):
	"""Act vs Budget for the single displayed month."""
	budget_key = month_key + "_budget"
	for row in rows or []:
		if not row:
			continue
		act = flt(row.get(month_key, 0))
		bud = flt(row.get(budget_key, 0))
		row["monthly_var_amount"] = act - bud
		row["monthly_var_percent"] = flt((row["monthly_var_amount"] / bud) * 100, 2) if bud else 0


def merge_prior_year(income, expense, filters, period_list, accounts_list, month_key):
	try:
		prev_fy = get_fiscal_year(
			add_years(period_list[0].year_start_date, -1),
			company=filters.company,
			as_dict=1,
		)
	except Exception:
		return

	prev_filters = frappe._dict(filters.copy())
	prev_filters.from_fiscal_year = prev_fy.name
	prev_filters.to_fiscal_year = prev_fy.name
	prev_filters.year = prev_fy.name

	prev_period_list = get_period_list(
		prev_filters.from_fiscal_year,
		prev_filters.to_fiscal_year,
		None,
		None,
		"Fiscal Year",
		filters.periodicity,
		company=filters.company,
	)
	to_m = list(MONTH_NAMES).index(filters.to_month)
	prev_period_list = [p for p in prev_period_list if p.to_date.month <= to_m]

	prev_income, prev_expense = fetch_pl(prev_filters, prev_period_list, accounts_list)
	_merge_prior_into(income, prev_income, prev_period_list)
	_merge_prior_into(expense, prev_expense, prev_period_list)


def _merge_prior_into(current_rows, prior_rows, prev_period_list):
	prior_map = {}
	prior_month_map = {}
	last_key = prev_period_list[-1].key if prev_period_list else None

	for row in prior_rows or []:
		key = row.get("account_origin") or row.get("account")
		if not key:
			key = row.get("account_name") or row.get("account")
		if not key:
			continue
		total = sum(flt(row.get(p.key, 0)) for p in prev_period_list)
		month_val = flt(row.get(last_key, 0)) if last_key else 0
		prior_map[key] = total
		prior_month_map[key] = month_val
		name = (row.get("account_name") or "").lower()
		if "total" in name and ("income" in name or "expense" in name or "revenue" in name):
			prior_map["__total__" + name] = total
			prior_month_map["__total__" + name] = month_val

	for row in current_rows or []:
		key = row.get("account_origin") or row.get("account")
		name = (row.get("account_name") or "").lower()
		if "total" in name and ("income" in name or "expense" in name or "revenue" in name):
			prior = flt(prior_map.get("__total__" + name, prior_map.get(key, 0)))
			prior_m = flt(prior_month_map.get("__total__" + name, prior_month_map.get(key, 0)))
		else:
			prior = flt(prior_map.get(key, 0))
			prior_m = flt(prior_month_map.get(key, 0))

		# YTD prior
		row["prior_total"] = prior
		actual_ytd = flt(row.get("total_actual", 0))
		row["prior_var_amount"] = actual_ytd - prior
		row["prior_var_percent"] = flt((row["prior_var_amount"] / prior) * 100, 2) if prior else 0

		# Monthly prior (same month last year)
		row["prior_month"] = prior_m
		# month actual is already on row under period key; var filled later if needed
		# Use displayed month from total path: caller sets via fill after knowing month_key
		# Store prior_month only; monthly prior var computed in fill_prior_month_var


def fill_prior_month_var(rows, month_key):
	for row in rows or []:
		if not row:
			continue
		act = flt(row.get(month_key, 0))
		prior_m = flt(row.get("prior_month", 0))
		row["prior_month_var_amount"] = act - prior_m
		row["prior_month_var_percent"] = (
			flt((row["prior_month_var_amount"] / prior_m) * 100, 2) if prior_m else 0
		)





def restructure_for_display(income, expense, filters, period_list, month_key):
	company = filters.company
	currency = (
		filters.get("presentation_currency")
		or frappe.get_cached_value("Company", filters.company, "default_currency")
	)

	account_parent_map = _build_account_parent_map(company)

	direct_income_parent = _get_group_account(company, "Direct Income")
	indirect_income_parent = _get_group_account(company, "Indirect Income")
	cogs_parent = _get_group_account(company, "Cost of Goods Sold")
	cos_parent = _get_group_account(company, "Cost of Sales")
	opex_parent = _get_group_account(company, "Expenses-Operating")
	depreciation_parent = _get_group_account(company, "Depreciation")
	finance_parent = _get_group_account(company, "Finance Expenses")
	tax_parent = _get_group_account_like(company, "Tax Expenses")

	def _classify_income(row):
		acc = row.get("account") or ""
		if _is_under(acc, direct_income_parent, account_parent_map):
			return "direct"
		if _is_under(acc, indirect_income_parent, account_parent_map):
			return "indirect"
		return "direct"

	def _classify_expense(row):
		acc = row.get("account") or ""
		if _is_under(acc, cogs_parent, account_parent_map):
			return "cogs"
		if _is_under(acc, cos_parent, account_parent_map):
			return "cos"
		if _is_under(acc, opex_parent, account_parent_map):
			return "opex"
		if _is_under(acc, depreciation_parent, account_parent_map):
			return "depreciation"
		if _is_under(acc, finance_parent, account_parent_map):
			return "finance"
		if tax_parent and _is_under(acc, tax_parent, account_parent_map):
			return "tax"
		return "opex"

	direct_rows = []
	indirect_rows = []
	cogs_rows = []
	cos_rows = []
	opex_rows = []
	depreciation_rows = []
	finance_rows = []
	tax_rows = []

	for row in income or []:
		if not row:
			continue
		if row.get("account_name") and "total" in (row.get("account_name") or "").lower():
			continue
		if row.get("is_group"):
			continue
		cat = _classify_income(row)
		if cat == "direct":
			direct_rows.append(row)
		else:
			indirect_rows.append(row)

	for row in expense or []:
		if not row:
			continue
		if row.get("account_name") and "total" in (row.get("account_name") or "").lower():
			continue
		if row.get("is_group"):
			continue
		cat = _classify_expense(row)
		if cat == "cogs":
			cogs_rows.append(row)
		elif cat == "cos":
			cos_rows.append(row)
		elif cat == "opex":
			opex_rows.append(row)
		elif cat == "depreciation":
			depreciation_rows.append(row)
		elif cat == "finance":
			finance_rows.append(row)
		elif cat == "tax":
			tax_rows.append(row)

	opex_grouped = _group_opex_by_parent(opex_rows, account_parent_map, company)

	value_keys = [p.key for p in period_list] + [p.key + "_budget" for p in period_list] + [
		"total_actual", "budget_ytd", "variance_amount", "variance_percent",
		"monthly_var_amount", "monthly_var_percent",
		"prior_total", "prior_var_amount", "prior_var_percent",
		"prior_month", "prior_month_var_amount", "prior_month_var_percent",
	]

	def _sum_rows(rows):
		result = {}
		for k in value_keys:
			result[k] = sum(flt(r.get(k, 0)) for r in rows)
		return result

	revenue_total = _sum_rows(direct_rows)
	cogs_total_vals = _sum_rows(cogs_rows)
	cos_total_vals = _sum_rows(cos_rows)
	cogs_cos_total = _sum_rows(cogs_rows + cos_rows)
	other_income_total = _sum_rows(indirect_rows)
	opex_total = _sum_rows(opex_rows)
	depreciation_total = _sum_rows(depreciation_rows)
	finance_total = _sum_rows(finance_rows)
	tax_total = _sum_rows(tax_rows)

	def _negate(vals):
		return {k: -flt(v) if k not in ("variance_percent", "monthly_var_percent", "prior_var_percent", "prior_month_var_percent") else v for k, v in vals.items()}

	gross_profit = {}
	for k in value_keys:
		gross_profit[k] = flt(revenue_total.get(k, 0)) - flt(cogs_cos_total.get(k, 0))

	_recalc_variances(gross_profit, month_key)

	ebitda = {}
	for k in value_keys:
		ebitda[k] = flt(gross_profit.get(k, 0)) + flt(other_income_total.get(k, 0)) - flt(opex_total.get(k, 0))
	_recalc_variances(ebitda, month_key)

	ebit = {}
	for k in value_keys:
		ebit[k] = flt(ebitda.get(k, 0)) - flt(depreciation_total.get(k, 0))
	_recalc_variances(ebit, month_key)

	pbt = {}
	for k in value_keys:
		pbt[k] = flt(ebit.get(k, 0)) - flt(finance_total.get(k, 0))
	_recalc_variances(pbt, month_key)

	pat = {}
	for k in value_keys:
		pat[k] = flt(pbt.get(k, 0)) - flt(tax_total.get(k, 0))
	_recalc_variances(pat, month_key)

	def _make_row(label, vals, bold=False, is_margin=False):
		row = frappe._dict({
			"account_name": label,
			"account": label,
			"currency": currency,
		})
		if bold:
			row["is_bold"] = 1
		if is_margin:
			row["is_margin"] = 1
		row.update(vals)
		return row

	def _margin_vals(numerator_vals, denominator_vals):
		result = {}
		for k in value_keys:
			num = flt(numerator_vals.get(k, 0))
			den = flt(denominator_vals.get(k, 0))
			if "percent" in k:
				result[k] = 0
			else:
				result[k] = flt((num / den) * 100, 2) if den else 0
		return result

	def _label_row(row, show_number=False):
		if row.get("is_group") and not show_number:
			row["account_name"] = remove_account_number(row.get("account_name") or "")
			if frappe.flags.in_export:
				row["account"] = row["account_name"]
		return row

	show_num = filters.get("show_number_group")
	data = []

	data.append(_make_row("Revenue", {}, bold=True))
	for r in direct_rows:
		if not r.get("is_group"):
			_label_row(r, show_num)
			r["indent"] = 1
			data.append(r)

	data.append({})

	cogs_cos_display = _sum_rows(cogs_rows + cos_rows)
	data.append(_make_row("COGS / Cost of Sales", cogs_cos_display, bold=True))

	cogs_display = _sum_rows(cogs_rows)
	data.append(_make_row("Cost of Goods Sold", cogs_display))
	for r in cogs_rows:
		if not r.get("is_group"):
			_label_row(r, show_num)
			r["indent"] = 2
			data.append(r)

	cos_display = _sum_rows(cos_rows)
	data.append(_make_row("Cost of Sales", cos_display))
	for r in cos_rows:
		if not r.get("is_group"):
			_label_row(r, show_num)
			r["indent"] = 2
			data.append(r)

	data.append({})

	data.append(_make_row("Gross Profit", gross_profit, bold=True))
	gp_margin = _margin_vals(gross_profit, revenue_total)
	data.append(_make_row("Gross Profit Margin", gp_margin, bold=True, is_margin=True))

	data.append(_make_row("Other income", {}, bold=True))
	for r in indirect_rows:
		if not r.get("is_group"):
			has_value = any(flt(r.get(k, 0)) for k in value_keys if "percent" not in k)
			if not has_value:
				continue
			_label_row(r, show_num)
			r["account_name"] = _friendly_other_income_name(r.get("account_name") or "")
			if frappe.flags.in_export:
				r["account"] = r["account_name"]
			r["indent"] = 1
			data.append(r)

	other_income_display = dict(other_income_total)
	data.append(_make_row("Total other income", other_income_display, bold=True))

	data.append(_make_row("Operating Expenses", {}, bold=True))
	for acc_num, label in OPEX_SUB_GROUPS:
		sub_rows = opex_grouped.get(acc_num, [])
		sub_total = _sum_rows(sub_rows)
		data.append(_make_row(label, sub_total))
		for r in sub_rows:
			if not r.get("is_group"):
				_label_row(r, show_num)
				r["indent"] = 2
				data.append(r)

	opex_display = dict(opex_total)
	data.append(_make_row("Total Operating Expenses", opex_display, bold=True))

	data.append({})

	data.append(_make_row("EBITDA", ebitda, bold=True))
	ebitda_margin = _margin_vals(ebitda, revenue_total)
	data.append(_make_row("EBITDA Margin", ebitda_margin, bold=True, is_margin=True))

	data.append({})

	dep_display = dict(depreciation_total)
	data.append(_make_row("Depreciation", dep_display, bold=True))

	data.append({})

	data.append(_make_row("EBIT", ebit, bold=True))
	ebit_margin = _margin_vals(ebit, revenue_total)
	data.append(_make_row("EBIT Margin", ebit_margin, bold=True, is_margin=True))

	data.append({})

	fin_display = dict(finance_total)
	data.append(_make_row("Finance Expenses", fin_display, bold=True))

	data.append({})

	data.append(_make_row("Profit / (Loss) before tax", pbt, bold=True))

	data.append({})

	tax_display = dict(tax_total)
	data.append(_make_row("Taxation", tax_display, bold=True))

	data.append({})

	data.append(_make_row("Profit / (Loss) after tax", pat, bold=True))
	npat_margin = _margin_vals(pat, revenue_total)
	data.append(_make_row("NPAT Margin", npat_margin, bold=True, is_margin=True))

	return data


def _recalc_variances(vals, month_key):
	vals["variance_amount"] = flt(vals.get("total_actual", 0)) - flt(vals.get("budget_ytd", 0))
	bud = flt(vals.get("budget_ytd", 0))
	vals["variance_percent"] = flt((vals["variance_amount"] / bud) * 100, 2) if bud else 0

	vals["prior_var_amount"] = flt(vals.get("total_actual", 0)) - flt(vals.get("prior_total", 0))
	prior = flt(vals.get("prior_total", 0))
	vals["prior_var_percent"] = flt((vals["prior_var_amount"] / prior) * 100, 2) if prior else 0

	act_m = flt(vals.get(month_key, 0))
	bud_m = flt(vals.get(month_key + "_budget", 0))
	prior_m = flt(vals.get("prior_month", 0))

	vals["monthly_var_amount"] = act_m - bud_m
	vals["monthly_var_percent"] = flt((vals["monthly_var_amount"] / bud_m) * 100, 2) if bud_m else 0

	vals["prior_month_var_amount"] = act_m - prior_m
	vals["prior_month_var_percent"] = flt((vals["prior_month_var_amount"] / prior_m) * 100, 2) if prior_m else 0


def _negate_for_display(vals):
	for k in list(vals.keys()):
		if "percent" in k:
			continue
		vals[k] = -flt(vals.get(k, 0))


def _negate_for_display_row(row):
	skip_keys = {"account_name", "account", "currency", "account_origin", "indent",
		"parent_account", "is_group", "is_bold", "is_margin", "warn_if_negative",
		"ratio_row", "profit_data", "acc_code", "year_start_date", "year_end_date"}
	for k in list(row.keys()):
		if k in skip_keys:
			continue
		if "percent" in k:
			continue
		val = row.get(k)
		if isinstance(val, (int, float)):
			row[k] = -val


def _friendly_other_income_name(name):
	mapping = {
		"Interest Income": "Interest income",
		"Miscellaneous Income": "Miscellaneous income",
		"Export Electricity Income": "Export electricity income",
		"Government Grants": "Government grants",
		"Amortisation Deferred Income": "Amortisation of deferred income",
		"Gain on Disposal of FA": "Gain on disposal of Fixed Assets",
	}
	clean = remove_account_number(name).strip()
	return mapping.get(clean, clean)


def _build_account_parent_map(company):
	accounts = frappe.db.sql(
		"""SELECT name, parent_account FROM `tabAccount` WHERE company=%s""",
		company,
		as_dict=1,
	)
	return {a.name: a.parent_account for a in accounts}


def _get_group_account(company, account_name):
	return frappe.db.get_value(
		"Account",
		{"account_name": account_name, "company": company, "is_group": 1},
		"name",
	)


def _get_group_account_like(company, pattern):
	result = frappe.db.sql(
		"""SELECT name FROM `tabAccount`
		WHERE company=%s AND is_group=1 AND account_name LIKE %s LIMIT 1""",
		(company, "%" + pattern + "%"),
	)
	return result[0][0] if result else None


def _is_under(account, parent, parent_map):
	if not parent:
		return False
	if account == parent:
		return True
	visited = set()
	current = account
	while current and current not in visited:
		visited.add(current)
		if current == parent:
			return True
		current = parent_map.get(current)
	return False


def _group_opex_by_parent(opex_rows, account_parent_map, company):
	group_accounts = {}
	for acc_num, label in OPEX_SUB_GROUPS:
		name = frappe.db.get_value(
			"Account",
			{"account_number": acc_num, "company": company, "is_group": 1},
			"name",
		)
		if name:
			group_accounts[acc_num] = name

	result = {acc_num: [] for acc_num, _ in OPEX_SUB_GROUPS}

	for row in opex_rows:
		acc = row.get("account") or ""
		placed = False
		for acc_num, _ in OPEX_SUB_GROUPS:
			parent_name = group_accounts.get(acc_num)
			if parent_name and _is_under(acc, parent_name, account_parent_map):
				result[acc_num].append(row)
				placed = True
				break
		if not placed:
			result[OPEX_SUB_GROUPS[-1][0]].append(row)

	return result


def get_report_columns(filters, period_list):
	currency = filters.get("presentation_currency") or frappe.get_cached_value(
		"Company", filters.company, "default_currency"
	)
	columns = [
		{
			"fieldname": "account",
			"label": _("Account"),
			"fieldtype": "Link",
			"options": "Account",
			"width": 300,
		},
		{
			"fieldname": "acc_code",
			"label": _("Acc. Code"),
			"fieldtype": "Data",
			"width": 110,
		},
	]

	year_str = str(filters.year)[-2:]
	prev_year_str = str(cint(filters.year) - 1)[-2:]
	month_abbr = (filters.to_month or "")[:3]

	for period in period_list:
		# Act May 26
		columns.append(
			{
				"fieldname": period.key,
				"label": _("Act {0} '{1}").format(month_abbr, year_str),
				"fieldtype": "Currency",
				"options": "currency",
				"width": 120,
			}
		)
		# Budget May 26
		columns.append(
			{
				"fieldname": period.key + "_budget",
				"label": _("Budget {0} '{1}").format(month_abbr, year_str),
				"fieldtype": "Currency",
				"options": "currency",
				"width": 130,
			}
		)

	columns.extend(
		[
			_cur("monthly_var_amount", _("Act vs Budget (Month) ($)"), width=200),
			_pct("monthly_var_percent", _("Act vs Budget (Month) (%)"), width=200),
			_cur("prior_month", _("Act {0} '{1}").format(month_abbr, prev_year_str)),
			_cur(
				"prior_month_var_amount",
				_("Act {0} '{1} vs Act {0} '{2} ($)").format(month_abbr, year_str, prev_year_str),
				width=230,
			),
			_pct(
				"prior_month_var_percent",
				_("Act {0} '{1} vs Act {0} '{2} (%)").format(month_abbr, year_str, prev_year_str),
				width=230,
			),
			_cur("total_actual", _("Actual YTD"), width=140),
			_cur("budget_ytd", _("Budget YTD"), width=140),
			_cur("variance_amount", _("Act vs Budget Yearly ($)"), width=200),
			_pct("variance_percent", _("Act vs Budget Yearly (%)"), width=200),
			_cur("prior_total", _("Act '{0}").format(prev_year_str)),
			_cur(
				"prior_var_amount",
				_("Act YTD '{0} vs Act YTD '{1} ($)").format(year_str, prev_year_str),
				width=230,
			),
			_pct(
				"prior_var_percent",
				_("Act YTD '{0} vs Act YTD '{1} (%)").format(year_str, prev_year_str),
				width=230,
			),
		]
	)

	return columns


def _cur(fieldname, label, width=140, hidden=0):
	col = {
		"fieldname": fieldname,
		"label": label,
		"fieldtype": "Currency",
		"options": "currency",
		"width": width,
	}
	if hidden:
		col["hidden"] = 1
	return col


def _pct(fieldname, label, width=120, hidden=0):
	col = {
		"fieldname": fieldname,
		"label": label,
		"fieldtype": "Percent",
		"width": width,
	}
	if hidden:
		col["hidden"] = 1
	return col


def add_borders(report_name, xlsx_file):
	import re
	from io import BytesIO

	from openpyxl import load_workbook
	from openpyxl.styles import Border, Side

	stream = BytesIO(xlsx_file.getvalue())
	wb = load_workbook(stream)
	thick = Side(style="thick")

	# Each box: (start pattern for first column, end pattern for last column)
	boxes = [
		(re.compile(r"^Act .+'\d+$"), re.compile(r"vs Act .+\(%\)$")),
		(re.compile(r"^Actual YTD$"), re.compile(r"vs Act YTD .+\(%\)$")),
	]

	def find_box(ws, start_re, end_re):
		header_row = min_col = max_col = None
		for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
			for cell in row:
				if not isinstance(cell.value, str):
					continue
				if min_col is None and start_re.match(cell.value):
					header_row = cell.row
					min_col = cell.column
				elif min_col is not None and max_col is None and end_re.search(cell.value):
					max_col = cell.column
			if header_row and min_col and max_col:
				break
		return header_row, min_col, max_col

	def find_last_row(ws, header_row, label="NPAT Margin"):
		for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=1, max_col=1):
			if row[0].value == label:
				return row[0].row
		return ws.max_row

	for ws in wb.worksheets:
		for start_re, end_re in boxes:
			header_row, min_col, max_col = find_box(ws, start_re, end_re)
			if not (header_row and min_col and max_col):
				continue

			max_row = find_last_row(ws, header_row)

			for row_idx in range(header_row, max_row + 1):
				for col_idx in range(min_col, max_col + 1):
					cell = ws.cell(row=row_idx, column=col_idx)
					cell.border = Border(
						top=thick if row_idx == header_row else cell.border.top,
						bottom=thick if row_idx == max_row else cell.border.bottom,
						left=thick if col_idx == min_col else cell.border.left,
						right=thick if col_idx == max_col else cell.border.right,
					)

	output_stream = BytesIO()
	wb.save(output_stream)
	output_stream.seek(0)

	from frappe.utils import get_datetime, now

	date_str_title = " " + get_datetime(now()).strftime("%y%m%d%H%M%S")
	frappe.response["filename"] = report_name + date_str_title + ".xlsx"
	frappe.response["filecontent"] = output_stream.getvalue()
	frappe.response["type"] = "binary"
