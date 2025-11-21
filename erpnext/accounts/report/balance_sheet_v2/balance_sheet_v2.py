# Copyright (c) 2015, Frappe Technologies Pvt. Ltd. and Contributors
# License: GNU General Public License v3. See license.txt


import frappe
from frappe import _
from frappe.utils import cint, flt, cstr

from erpnext.accounts.report.financial_statements import (
	get_columns,
	get_data,
	get_filtered_list_for_consolidated_report,
	get_period_list,
)

from erpnext.accounts.utils import remove_account_number
from erpnext.accounts.report.utils import convert_wrap_report_data


def execute(filters=None):
	# Ensure filters behaves like a dict
	filters = frappe._dict(filters or {})

	period_list = get_period_list(
		filters.from_fiscal_year,
		filters.to_fiscal_year,
		filters.period_start_date,
		filters.period_end_date,
		filters.filter_based_on,
		filters.periodicity,
		company=filters.company,
		month=filters.month,
		to_month=filters.to_month,
	)

	# If user asks for Monthly Net (non-accumulated), force non-accumulated values
	if filters.get("periodicity") == "Monthly" and filters.get("monthly_net"):
		filters.accumulated_values = 0
	# Default to accumulated across all periodicities if not explicitly set
	if filters.get("accumulated_values") is None:
		filters.accumulated_values = 1

	currency = filters.presentation_currency or frappe.get_cached_value(
		"Company", filters.company, "default_currency"
	)

	ignore_closing_entries = filters.ignore_closing_entries

	asset = get_data(
		filters.company,
		"Asset",
		"Debit",
		period_list,
		only_current_fiscal_year=False,
		filters=filters,
		accumulated_values=filters.accumulated_values,
		ignore_closing_entries=ignore_closing_entries
	)

	liability = get_data(
		filters.company,
		"Liability",
		"Credit",
		period_list,
		only_current_fiscal_year=False,
		filters=filters,
		accumulated_values=filters.accumulated_values,
		ignore_closing_entries=ignore_closing_entries
	)

	equity = get_data(
		filters.company,
		"Equity",
		"Credit",
		period_list,
		only_current_fiscal_year=False,
		filters=filters,
		accumulated_values=filters.accumulated_values,
		ignore_closing_entries=ignore_closing_entries
	)

	provisional_profit_loss, total_credit = get_provisional_profit_loss(
		asset, liability, equity, period_list, filters.company, currency
	)

	message, opening_balance = check_opening_balance(asset, liability, equity)

	temp_data = []
	temp_data.extend(asset or [])
	temp_data.extend(liability or [])
	temp_data.extend(equity or [])

	data = []
	if filters.show_number_group:
		data = temp_data
	else:
		for d in temp_data:
			if d.get("is_group"):
				d['account_name'] = remove_account_number(d['account_name'])
				if frappe.flags.in_export:
					d['account_origin'] = cstr(d['account'])
					d['account'] = d['account_name']
					if not d.get('parent_account'):
						for key, val in d.items():
							if key not in ['account', 'account_name']:
								d[key] = None

			data.append(d)

	if opening_balance and round(opening_balance, 2) != 0:
		unclosed = {
			"account_name": "'" + _("Unclosed Fiscal Years Profit / Loss (Credit)") + "'",
			"account": "'" + _("Unclosed Fiscal Years Profit / Loss (Credit)") + "'",
			"warn_if_negative": True,
			"currency": currency,
		}
		for period in period_list:
			unclosed[period.key] = opening_balance
			if provisional_profit_loss:
				provisional_profit_loss[period.key] = provisional_profit_loss[period.key] - opening_balance

		unclosed["total"] = opening_balance
		data.append(unclosed)

	if provisional_profit_loss:
		data.append(provisional_profit_loss)
	if total_credit:
		data.append(total_credit)

	columns = get_columns(
		filters.periodicity, period_list, filters.accumulated_values, company=filters.company
	)

	chart = get_chart_data(filters, columns, asset, liability, equity)

	report_summary = get_report_summary(
		period_list, asset, liability, equity, provisional_profit_loss, total_credit, currency, filters
	)

	# if frappe.flags.in_export:
	# 	convert_wrap_report_data(columns, data, precision=2)

	return columns, data, message, chart, report_summary


def get_provisional_profit_loss(
	asset, liability, equity, period_list, company, currency=None, consolidated=False
):
	provisional_profit_loss = {}
	total_row = {}
	if asset and (liability or equity):
		total = total_row_total = 0
		currency = currency or frappe.get_cached_value("Company", company, "default_currency")
		total_row = {
			"account_name": "'" + _("Total (Credit)") + "'",
			"account": "'" + _("Total (Credit)") + "'",
			"warn_if_negative": True,
			"currency": currency,
		}
		has_value = False

		for period in period_list:
			key = period if consolidated else period.key
			effective_liability = 0.0
			if liability:
				effective_liability += flt(liability[-2].get(key))
			if equity:
				effective_liability += flt(equity[-2].get(key))

			provisional_profit_loss[key] = flt(asset[-2].get(key)) - effective_liability
			total_row[key] = effective_liability + provisional_profit_loss[key]

			if provisional_profit_loss[key]:
				has_value = True

			provisional_profit_loss["total"] = flt(total)
			total += flt(provisional_profit_loss[key])

			total_row_total += flt(total_row[key])
			total_row["total"] = total_row_total

		if has_value:
			provisional_profit_loss.update(
				{
					"account_name": "'" + _("Profit / (Loss) for the Year") + "'",
					"account": "'" + _("Profit / (Loss) for the Year") + "'",
					"warn_if_negative": True,
					"currency": currency,
				}
			)

	return provisional_profit_loss, total_row


def check_opening_balance(asset, liability, equity):
	# Check if previous year balance sheet closed
	opening_balance = 0
	float_precision = cint(frappe.db.get_default("float_precision")) or 2
	if asset:
		opening_balance = flt(asset[-1].get("opening_balance", 0), float_precision)
	if liability:
		opening_balance -= flt(liability[-1].get("opening_balance", 0), float_precision)
	if equity:
		opening_balance -= flt(equity[-1].get("opening_balance", 0), float_precision)

	opening_balance = flt(opening_balance, float_precision)
	if opening_balance:
		return _("Previous Financial Year is not closed"), opening_balance
	return None, None


def get_report_summary(
	period_list,
	asset,
	liability,
	equity,
	provisional_profit_loss,
	total_credit,
	currency,
	filters,
	consolidated=False,
):

	net_asset, net_liability, net_equity, net_provisional_profit_loss = 0.0, 0.0, 0.0, 0.0

	if filters.get("accumulated_values"):
		period_list = [period_list[-1]]

	# from consolidated financial statement
	if filters.get("accumulated_in_group_company"):
		period_list = get_filtered_list_for_consolidated_report(filters, period_list)

	for period in period_list:
		key = period if consolidated else period.key
		if asset:
			net_asset += asset[-2].get(key)
		if liability:
			net_liability += liability[-2].get(key)
		if equity:
			net_equity += equity[-2].get(key)
		if provisional_profit_loss:
			net_provisional_profit_loss += provisional_profit_loss.get(key)

	return [
		{"value": net_asset, "label": _("Total Asset"), "datatype": "Currency", "currency": currency},
		{
			"value": net_liability,
			"label": _("Total Liability"),
			"datatype": "Currency",
			"currency": currency,
		},
		{"value": net_equity, "label": _("Total Equity"), "datatype": "Currency", "currency": currency},
		{
			"value": net_provisional_profit_loss,
			"label": _("Profit / (Loss) for the Year"),
			"indicator": "Green" if net_provisional_profit_loss > 0 else "Red",
			"datatype": "Currency",
			"currency": currency,
		},
	]


def get_chart_data(filters, columns, asset, liability, equity):
	labels = [d.get("label") for d in columns[2:]]

	asset_data, liability_data, equity_data = [], [], []

	for p in columns[2:]:
		if asset:
			asset_data.append(asset[-2].get(p.get("fieldname")))
		if liability:
			liability_data.append(liability[-2].get(p.get("fieldname")))
		if equity:
			equity_data.append(equity[-2].get(p.get("fieldname")))

	datasets = []
	if asset_data:
		datasets.append({"name": _("Assets"), "values": asset_data})
	if liability_data:
		datasets.append({"name": _("Liabilities"), "values": liability_data})
	if equity_data:
		datasets.append({"name": _("Equity"), "values": equity_data})

	chart = {"data": {"labels": labels, "datasets": datasets}}

	if not cint(filters.accumulated_values):
		chart["type"] = "bar"
	else:
		chart["type"] = "line"

	return chart

from frappe.utils.xlsxutils import make_xlsx
from openpyxl import load_workbook
from io import BytesIO
import json
from openpyxl.styles import Font
from frappe.utils import now_datetime
def add_formulas(report_name, xlsx_file):
	stream = BytesIO(xlsx_file.getvalue())
	wb = load_workbook(stream)
	ws = wb.active

	group_col = "Z"
	ws.column_dimensions[group_col].hidden = True

	col_use = list(ws.column_dimensions.keys())

	level = 0 
	start_row = 0
	for row in range(2, ws.max_row + 1): 
		account_txt = cstr(ws[f"A{row}"].value) 
		leading_spaces = len(account_txt) - len(account_txt.lstrip(" ")) 
		account = cstr(ws[f"A{row}"].value).strip() 

		if account in ['Assets', 'Income']: 
			start_row = row
			break

	rows = build_rows(ws, start_row=start_row)

	hier = compute_child_range_rows(rows)
	flat_list = extract_nodes(hier)

	assets_total_row = 0
	liability_total_row = 0
	total_row = {}
	for d in flat_list:
		account = d['account']
		row = d['row']
		# Balance Sheet
		if account == "Total Asset (Debit)":
			total_row.setdefault("Assets", row)
		elif account == "Total Liability (Credit)":
			total_row.setdefault("Liabilities", row)
		elif account == "Total Equity (Credit)":
			total_row.setdefault("Equity", row)

		# Profit & Loss
		elif account == "Total Income (Credit)":
			total_row.setdefault("Income", row)
		elif account == "Total Expense (Debit)":
			total_row.setdefault("Expenses", row)

	for d in flat_list:
		is_group = d.get("group_flag")
		account = d['account']
		row = d['row']
		lft=d['lft_row']
		rgt=d['rgt_row']
		ws[f'{group_col}{row}'] = 2 if is_group else 1

		for cell in ws[row]:
			cell.font = Font(bold=bool(is_group))
		
		if is_group and lft:
			for col in col_use:
				if col not in ['A', 'B', group_col]:
					ws[f"{col}{row}"] = f"=SUMIF({group_col}{lft}:{group_col}{rgt},1,{col}{lft}:{col}{rgt})"

					# Hardcode formula
					if account in total_row:
						ws[f"{col}{total_row.get(account)}"] = f"=SUMIF({group_col}{lft}:{group_col}{rgt},1,{col}{lft}:{col}{rgt})"
		else:
			for col in col_use:
				if col not in ['A', 'B', group_col]:
					# Balance Sheet
					required_keys = ["Assets", "Liabilities", "Equity"]
					if all(k in total_row for k in required_keys):
						equity_total_row = total_row.get("Equity")
						liability_total_row = total_row.get("Liabilities")
						assets_total_row = total_row.get("Assets")
						if account == "'Profit / (Loss) for the Year'":
							ws[f"{col}{row}"] = f"={col}{assets_total_row} - ({col}{liability_total_row} + {col}{equity_total_row})"
						elif account == "'Total (Credit)'":
							ws[f"{col}{row}"] = f"={col}{liability_total_row} + {col}{equity_total_row}"
					
					# Profit & Loss
					required_keys = ["Income", "Expenses"]
					if all(k in total_row for k in required_keys):
						if account == "'Profit for the year'":
							income_total_row = total_row.get("Income")
							expense_total_row = total_row.get("Expenses")
							ws[f"{col}{row}"] = f"={col}{income_total_row} - {col}{expense_total_row}"

	output_stream = BytesIO()
	wb.save(output_stream)
	output_stream.seek(0)

	now = now_datetime()
	date_str_title = now.strftime("%y%m%d_%H%M%S")

	frappe.response["filename"] = f"{report_name}_{date_str_title}.xlsx"
	frappe.response["filecontent"] = output_stream.getvalue()
	frappe.response["type"] = "binary"

import re
from collections import OrderedDict
def count_leading_spaces(s):
	if not isinstance(s, str):
		return 0
	return len(re.match(r"^[\s\xa0]*", s).group(0))

def get_level(account_txt, spaces_per_level=4):
	return count_leading_spaces(account_txt) // spaces_per_level

def is_child(cell_value: str) -> bool:
    if not isinstance(cell_value, str):
        return False
    s = cell_value.strip()
    return bool(s) and s[0].isdigit()

def build_rows(ws, start_row=2, start_after_label=None, spaces_per_level=4):
    """Ambil linear rows: [(row_idx, account, level, group_flag)]"""
    rows = []
    started = (start_after_label is None)

    for r in range(start_row, ws.max_row + 1):
        raw = cstr(ws[f"A{r}"].value or "")
        if not raw.strip():
            continue

        if not started:
            if raw.strip() == start_after_label:
                started = True
            else:
                continue

        level = get_level(raw, spaces_per_level)
        account = raw.strip()
        group_flag = 0 if is_child(account) else 1  # 1=group, 0=child
        rows.append((r, account, level, group_flag))

    return rows

def compute_child_range_rows(rows):
    """
    For each group, compute (lft_row, rgt_row) = the range of child rows.
    Returns: OrderedDict with key (account, row, group_flag, lft_row, rgt_row), value = children (tree)
    """
    # Build the tree structure using a stack (based on indentation level)
    stack = []  # elements: (level, key_tuple, children_dict_ref)
    roots = OrderedDict()

    # Keep an index mapping for convenience
    nodes = []  # [(i, row_idx, level, group_flag, key_ref_dict, parent_children_dict)]
    tmp_keys = []  # list of (account, row_idx, group_flag) – lft/rgt will be added later

    for i, (row_idx, account, level, group_flag) in enumerate(rows):
        key_tmp = (account, row_idx, group_flag)
        tmp_keys.append(key_tmp)
        node_children = OrderedDict()

        # Pop until parent level < current level
        while stack and stack[-1][0] >= level:
            stack.pop()

        # Attach to parent if exists, otherwise treat as root
        if stack:
            parent_children = stack[-1][2]
            parent_children[key_tmp] = node_children
        else:
            roots[key_tmp] = node_children

        # Push current node to stack
        stack.append((level, key_tmp, node_children))
        nodes.append((i, row_idx, level, group_flag, key_tmp, node_children))

    # Determine (lft_row, rgt_row) range for each group node
    # Final result replaces key_tmp → key_final (account, row, group_flag, lft_row, rgt_row)
    def child_block_range(i_parent):
        row_idx_p, level_p = rows[i_parent][0], rows[i_parent][2]
        # Find the first index j where level <= parent level
        j = i_parent + 1
        while j < len(rows) and rows[j][2] > level_p:
            j += 1
        # Child range = i_parent+1 .. j-1
        if j - (i_parent + 1) >= 1:
            lft_row = rows[i_parent + 1][0]
            rgt_row = rows[j - 1][0]
            return lft_row, rgt_row
        return None, None

    # Recursive function to materialize the final structure with (lft_row, rgt_row)
    def materialize(children_dict, start_index_lookup):
        out = OrderedDict()
        for key_tmp, ch in children_dict.items():
            account, row_idx, group_flag = key_tmp
            # Find the row index of this node
            # (can use dict lookup for performance, but linear scan is fine for report-sized data)
            i = next(i for i, (r, a, lvl, gf) in enumerate(rows)
                     if r == row_idx and a == account and gf == group_flag)

            if group_flag == 1:
                lft_row, rgt_row = child_block_range(i)
            else:
                lft_row = rgt_row = None

            key_final = (account, row_idx, group_flag, lft_row, rgt_row)
            out[key_final] = materialize(ch, start_index_lookup)
        return out

    # Build the final ordered structure
    result = OrderedDict()
    result = materialize(roots, {r[0]: idx for idx, r in enumerate(rows)})
    return result


def extract_nodes(flat_tree, parent=None, result=None):
    """
    Ubah struktur tree hasil compute_child_range_rows menjadi list datar.
    Setiap elemen hasil = (account, row, group_flag, lft_row, rgt_row, parent_row)
    """
    if result is None:
        result = []

    for (account, row, group_flag, lft_row, rgt_row), children in flat_tree.items():
        result.append({
            "account": account,
            "row": row,
            "group_flag": group_flag,
            "lft_row": lft_row,
            "rgt_row": rgt_row,
            "parent_row": parent[1] if parent else None
        })
        extract_nodes(children, (account, row, group_flag, lft_row, rgt_row), result)

    return result
