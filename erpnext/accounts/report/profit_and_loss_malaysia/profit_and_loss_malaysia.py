# Copyright (c) 2026, Frappe Technologies Pvt. Ltd. and contributors
# For license information, please see license.txt

import frappe
import re
import json
from urllib.parse import quote
from io import BytesIO
try:
	import openpyxl
except Exception:
	openpyxl = None
from frappe import _
from frappe.utils import flt, now_datetime, get_datetime, now, cint

from erpnext.accounts.report.financial_statements import (
	get_columns,
	get_data,
	get_filtered_list_for_consolidated_report,
	get_period_list,
)
from erpnext.accounts.utils import remove_account_number

COMPANY = "Greenphyto Tech Sdn Bhd"

COGS_GROUPS = ["Cost of Goods Sold", "Cost of Sales"]
OPERATING_GROUPS = ["Expenses-Operating"]
DEPRECIATION_GROUPS = ["Depreciation"]
FINANCE_GROUPS = ["Finance Expenses"]
TAX_GROUPS = ["Tax Expenses- GPL"]


def execute(filters=None):
	filters = frappe._dict(filters)

	fiscal_year = filters.get("fiscal_year") or filters.get("from_fiscal_year")
	filters.from_fiscal_year = fiscal_year
	filters.to_fiscal_year = fiscal_year
	filters.periodicity = "Monthly"
	filters.accumulated_values = 0
	filters.filter_based_on = "Fiscal Year"

	period_list = get_period_list(
		filters.from_fiscal_year,
		filters.to_fiscal_year,
		filters.period_start_date,
		filters.period_end_date,
		filters.filter_based_on,
		filters.periodicity,
		company=filters.company,
		month=filters.month,
		to_month=filters.to_month,
	)

	from frappe.utils import getdate, today
	current_date = getdate(today())
	fy_year = int(fiscal_year) if fiscal_year else current_date.year

	if fy_year >= current_date.year:
		current_month = current_date.month
		period_list = [p for p in period_list if getdate(p.to_date).month <= current_month]

	hide_if_empty = cint(filters.get("hide_if_empty", 1))

	income = get_data(
		filters.company,
		"Income",
		"Credit",
		period_list,
		filters=filters,
		accumulated_values=filters.accumulated_values,
		ignore_closing_entries=True,
		ignore_accumulated_values_for_fy=True,
		filter_zero_value=hide_if_empty,
	)

	expense = get_data(
		filters.company,
		"Expense",
		"Debit",
		period_list,
		filters=filters,
		accumulated_values=filters.accumulated_values,
		ignore_closing_entries=True,
		ignore_accumulated_values_for_fy=True,
		filter_zero_value=hide_if_empty,
	)

	data = build_malaysia_pnl(income, expense, period_list, filters)

	columns = get_columns(
		filters.periodicity, period_list, filters.accumulated_values, filters.company
	)

	# Replace "Total" column with "Year to-date"
	for col in columns:
		if col.get("fieldname") == "total":
			col["label"] = "Year to-date"

	# Add YTD values to data rows
	for row in data:
		ytd = 0
		for period in period_list:
			ytd += flt(row.get(period.key, 0))
		row["total"] = ytd

	chart = get_chart_data(filters, columns, data, period_list)

	currency = filters.presentation_currency or frappe.get_cached_value(
		"Company", filters.company, "default_currency"
	)
	report_summary = get_report_summary(data, period_list, currency, filters)

	return columns, data, None, chart, report_summary


def build_malaysia_pnl(income, expense, period_list, filters):
	data = []

	income_rows = income or []
	expense_rows = expense or []

	cogs_rows, operating_rows, depreciation_rows, finance_rows, tax_rows = classify_expense_rows(expense_rows)

	# === INCOME SECTION ===
	data.append(make_header_row("Income"))
	for row in income_rows:
		account_name = row.get("account_name", "")
		if not account_name or "total" in account_name.lower() or account_name == "Income":
			continue
		clean_row = normalize_indent(clean_account_row(row), base_indent=1)
		data.append(clean_row)

	revenue_totals = get_section_totals(income_rows, period_list)
	data.append(make_value_row("Revenue", revenue_totals, period_list, is_bold=True))
	data.append(make_empty_row())

	# === EXPENSES SECTION ===
	data.append(make_header_row("Expenses"))

	# COGS and Cost of Sales
	for row in cogs_rows:
		if "total" not in row.get("account_name", "").lower():
			clean_row = normalize_indent(clean_account_row(row), base_indent=1, min_indent=1)
			data.append(clean_row)

	cogs_cos_totals = get_rows_total(cogs_rows, period_list)
	neg_cogs = negate_values(cogs_cos_totals, period_list)
	data.append(make_value_row("COGS / Cost of Sales", neg_cogs, period_list, is_bold=True))
	data.append(make_empty_row())

	# === GROSS PROFIT ===
	gross_profit = {}
	for period in period_list:
		rev = flt(revenue_totals.get(period.key), 2)
		cogs = flt(cogs_cos_totals.get(period.key), 2)
		gross_profit[period.key] = rev - cogs

	data.append(make_value_row("Gross Profit", gross_profit, period_list, is_bold=True))

	# Gross Profit Margin
	gp_margin = {}
	for period in period_list:
		rev = flt(revenue_totals.get(period.key), 2)
		if rev:
			gp_margin[period.key] = flt(gross_profit[period.key] / rev, 6)
		else:
			gp_margin[period.key] = 0
	data.append(make_value_row("Gross Profit Margin", gp_margin, period_list, is_ratio=True))
	data.append(make_empty_row())

	# === OPERATING EXPENSES ===
	data.append(make_header_row("Operating Expenses"))
	for row in operating_rows:
		if "total" not in row.get("account_name", "").lower():
			clean_row = normalize_indent(clean_account_row(row), base_indent=0)
			data.append(clean_row)

	opex_totals = get_rows_total(operating_rows, period_list)
	neg_opex = negate_values(opex_totals, period_list)
	data.append(make_value_row("Total Operating Expense", neg_opex, period_list, is_bold=True))
	data.append(make_empty_row())

	# === EBITDA ===
	ebitda = {}
	for period in period_list:
		ebitda[period.key] = flt(gross_profit[period.key]) - flt(opex_totals.get(period.key))

	data.append(make_value_row("EBITDA", ebitda, period_list, is_bold=True))

	ebitda_margin = {}
	for period in period_list:
		rev = flt(revenue_totals.get(period.key), 2)
		if rev:
			ebitda_margin[period.key] = flt(ebitda[period.key] / rev, 6)
		else:
			ebitda_margin[period.key] = 0
	data.append(make_value_row("EBITDA Margin", ebitda_margin, period_list, is_ratio=True))
	data.append(make_empty_row())

	# === DEPRECIATION AND AMORTISATION ===
	data.append(make_header_row("Depreciation and Amortisation"))
	for row in depreciation_rows:
		if "total" not in row.get("account_name", "").lower():
			clean_row = normalize_indent(clean_account_row(row), base_indent=0, min_indent=2)
			data.append(clean_row)

	da_totals = get_rows_total(depreciation_rows, period_list)
	data.append(make_empty_row())

	# === EBIT ===
	ebit = {}
	for period in period_list:
		ebit[period.key] = flt(ebitda[period.key]) - flt(da_totals.get(period.key))

	data.append(make_value_row("EBIT", ebit, period_list, is_bold=True))

	ebit_margin = {}
	for period in period_list:
		rev = flt(revenue_totals.get(period.key), 2)
		if rev:
			ebit_margin[period.key] = flt(ebit[period.key] / rev, 6)
		else:
			ebit_margin[period.key] = 0
	data.append(make_value_row("EBIT Margin", ebit_margin, period_list, is_ratio=True))
	data.append(make_empty_row())

	# === INTEREST (FINANCE EXPENSES) ===
	data.append(make_header_row("Interest"))
	for row in finance_rows:
		if "total" not in row.get("account_name", "").lower():
			clean_row = normalize_indent(clean_account_row(row), base_indent=0, min_indent=2)
			data.append(clean_row)

	interest_totals = get_rows_total(finance_rows, period_list)
	data.append(make_empty_row())

	# === PROFIT / (LOSS) BEFORE TAX ===
	pbt = {}
	for period in period_list:
		pbt[period.key] = flt(ebit[period.key]) - flt(interest_totals.get(period.key))

	data.append(make_value_row("Profit / (Loss) before tax", pbt, period_list, is_bold=True))
	data.append(make_empty_row())

	# === TAXATION ===
	data.append(make_header_row("Taxation"))
	for row in tax_rows:
		if "total" not in row.get("account_name", "").lower():
			clean_row = normalize_indent(clean_account_row(row), base_indent=0, min_indent=2)
			data.append(clean_row)

	tax_totals = get_rows_total(tax_rows, period_list)
	data.append(make_empty_row())

	# === PROFIT / (LOSS) AFTER TAX ===
	pat = {}
	for period in period_list:
		pat[period.key] = flt(pbt[period.key]) - flt(tax_totals.get(period.key))

	data.append(make_value_row("Profit / (Loss) after tax", pat, period_list, is_bold=True))

	# NPAT Margin
	npat_margin = {}
	for period in period_list:
		rev = flt(revenue_totals.get(period.key), 2)
		if rev:
			npat_margin[period.key] = flt(pat[period.key] / rev, 6)
		else:
			npat_margin[period.key] = 0
	data.append(make_value_row("NPAT Margin", npat_margin, period_list, is_ratio=True))

	return data


def classify_expense_rows(expense_rows):
	cogs_rows = []
	operating_rows = []
	depreciation_rows = []
	finance_rows = []
	tax_rows = []

	current_section = None
	skip_names = ["Expenses", "Expenses-Other", "Total Expense (Debit)"]

	for row in expense_rows:
		account_name = row.get("account_name", "")

		if account_name in skip_names or not account_name:
			continue

		if account_name in COGS_GROUPS:
			current_section = "cogs"
		elif account_name in OPERATING_GROUPS:
			current_section = "operating"
			continue
		elif account_name in DEPRECIATION_GROUPS:
			current_section = "depreciation"
			continue
		elif account_name in FINANCE_GROUPS:
			current_section = "finance"
			continue
		elif account_name in TAX_GROUPS:
			current_section = "tax"
			continue

		if current_section == "cogs":
			cogs_rows.append(row)
		elif current_section == "operating":
			operating_rows.append(row)
		elif current_section == "depreciation":
			depreciation_rows.append(row)
		elif current_section == "finance":
			finance_rows.append(row)
		elif current_section == "tax":
			tax_rows.append(row)

	return cogs_rows, operating_rows, depreciation_rows, finance_rows, tax_rows


def get_section_totals(rows, period_list):
	totals = {}
	for period in period_list:
		totals[period.key] = 0

	if not rows:
		return totals

	for row in rows:
		account_name = row.get("account_name", "").lower()
		if "total" in account_name:
			for period in period_list:
				totals[period.key] = flt(row.get(period.key), 2)
			return totals

	return totals


def get_rows_total(rows, period_list):
	totals = {}
	for period in period_list:
		totals[period.key] = 0

	for row in rows:
		if row.get("is_group"):
			continue
		account_name = row.get("account_name", "").lower()
		if "total" in account_name:
			continue

		for period in period_list:
			totals[period.key] += flt(row.get(period.key), 2)

	return totals


def negate_values(values, period_list):
	neg = {}
	for period in period_list:
		neg[period.key] = -1 * flt(values.get(period.key), 2)
	return neg


def make_header_row(label):
	return {
		"account_name": label,
		"account": label,
		"is_group": 1,
		"indent": 0,
		"currency": "MYR",
	}


def make_value_row(label, values, period_list, is_bold=False, is_ratio=False):
	row = {
		"account_name": label,
		"account": label,
		"is_bold": is_bold,
		"indent": 0,
		"currency": "MYR",
	}
	if is_ratio:
		row["is_ratio"] = 1

	for period in period_list:
		row[period.key] = values.get(period.key, 0)

	return row


def make_empty_row():
	return {
		"account_name": "",
		"account": "",
		"currency": "MYR",
	}


def clean_account_row(row):
	cleaned = dict(row)
	account_name = cleaned.get("account_name", "")
	if cleaned.get("is_group"):
		cleaned["account_name"] = remove_account_number(account_name)
	return cleaned


def normalize_indent(row, base_indent=0, min_indent=1):
	original_indent = int(row.get("indent", 0))
	if original_indent > min_indent:
		row["indent"] = original_indent - min_indent + base_indent
	else:
		row["indent"] = base_indent
	return row


def get_chart_data(filters, columns, data, period_list):
	labels = [p.label for p in period_list]

	revenue_values = []
	expense_values = []
	profit_values = []

	for row in data:
		if row.get("account_name") == "Revenue":
			for p in period_list:
				revenue_values.append(flt(row.get(p.key)))
		elif row.get("account_name") == "Total Operating Expense":
			for p in period_list:
				expense_values.append(flt(row.get(p.key)))
		elif row.get("account_name") == "Profit / (Loss) after tax":
			for p in period_list:
				profit_values.append(flt(row.get(p.key)))

	datasets = []
	if revenue_values:
		datasets.append({"name": _("Revenue"), "values": revenue_values})
	if expense_values:
		datasets.append({"name": _("Operating Expense"), "values": expense_values})
	if profit_values:
		datasets.append({"name": _("Net Profit/Loss"), "values": profit_values})

	chart = {"data": {"labels": labels, "datasets": datasets}}

	if not cint(filters.accumulated_values):
		chart["type"] = "bar"
	else:
		chart["type"] = "line"

	chart["fieldtype"] = "Currency"
	return chart


def get_report_summary(data, period_list, currency, filters):
	revenue = 0
	total_expense = 0
	net_profit = 0

	for row in data:
		if row.get("account_name") == "Revenue":
			for p in period_list:
				revenue += flt(row.get(p.key))
		elif row.get("account_name") == "Total Operating Expense":
			for p in period_list:
				total_expense += flt(row.get(p.key))
		elif row.get("account_name") == "Profit / (Loss) after tax":
			for p in period_list:
				net_profit += flt(row.get(p.key))

	return [
		{"value": revenue, "label": _("Total Revenue"), "datatype": "Currency", "currency": currency},
		{"type": "separator", "value": "-"},
		{"value": total_expense, "label": _("Total Operating Expense"), "datatype": "Currency", "currency": currency},
		{"type": "separator", "value": "=", "color": "blue"},
		{
			"value": net_profit,
			"indicator": "Green" if net_profit > 0 else "Red",
			"label": _("Net Profit/Loss"),
			"datatype": "Currency",
			"currency": currency,
		},
	]


from frappe.desk.query_report import add_title_report, get_filters_data, build_xlsx_data
from openpyxl.utils import get_column_letter

def get_export_cost_center(report_name, filters):
	base_filters = frappe._dict(filters or {})

	cc_filters = {"is_group": 0}
	if base_filters.get("company"):
		cc_filters["company"] = base_filters.company

	cost_centers = [
		frappe._dict({"name": "All Cost Centers", "cost_center_name": "All Cost Centers"})
	]
	cost_centers += frappe.get_all(
		"Cost Center",
		filters=cc_filters,
		fields=["name", "cost_center_name"],
		order_by="name asc",
	) or []

	group_data = {}

	export_date = now()
	date_str = " "+get_datetime(export_date).strftime("%-d %B %y %H:%M:%S")
	title_report = add_title_report(report_name, filters)

	from erpnext.accounts.report.profit_and_loss_malaysia.profit_and_loss_malaysia import execute

	for idx, cc in enumerate(cost_centers):
		per_cc_filters = frappe._dict(base_filters.copy())
		if idx == 0:
			per_cc_filters["cost_center"] = []
		else:
			per_cc_filters["cost_center"] = [cc.name]

		filter_report = get_filters_data(per_cc_filters)
		columns, data, _, _, report_summary = execute(per_cc_filters)
		temp = frappe._dict({
			"columns": columns,
			"result": data
		})
		xlsx_data, column_widths = build_xlsx_data(temp, [], 1, ignore_visible_idx=1)
		header_data = title_report + filter_report + [["Export date", date_str]]
		xlsx_data = header_data + xlsx_data

		start_from = len(header_data) + 2
		bold_list = []
		for i, d in enumerate(header_data):
			if d and d[0] == "cost_center":
				bold_list.append(i+1)
			if d and d[0] == "Export date":
				bold_list.append(i+1)

		for i, d in enumerate(data):
			if i == 0:
				bold_list.append(start_from-1)
			if d.get("is_group") or d.get("is_bold") or d.get("bold"):
				idx = i + start_from
				bold_list.append(idx)

		group_data[cc.name] = {
			"label": cc.cost_center_name or cc.name,
			"columns": columns,
			"data": xlsx_data,
			"summary": report_summary,
			"column_widths": column_widths,
			"bold_list": bold_list
		}

	return group_data


def _sanitize_sheet_name(name):
	name = (name or "Sheet").strip()
	name = re.sub(r"[:\\/\?\*\[\]]", " ", name)
	if len(name) > 31:
		name = name[:31]
	return name or "Sheet"


@frappe.whitelist()
def export_with_cost_centers(report_name, filters=None, formula=False):
	if isinstance(filters, str):
		try:
			filters = frappe.parse_json(filters)
		except Exception:
			filters = {}

	formula = cint(formula)
	filters['show_all_cost_centers'] = 0
	group_data = get_export_cost_center(report_name, filters)

	if not openpyxl:
		frappe.throw("openpyxl is required to export XLSX")

	wb = openpyxl.Workbook(write_only=True)

	used_names = set()
	sheet_bold_map = {}
	right_cell = []

	for cc_name, payload in (group_data or {}).items():
		label = payload.get("label") or cc_name
		sheet_name = _sanitize_sheet_name(label)

		base = sheet_name
		idx = 1
		while sheet_name in used_names:
			suffix = f" {idx}"
			sheet_name = _sanitize_sheet_name((base[: (31 - len(suffix))]).rstrip() + suffix)
			idx += 1
		used_names.add(sheet_name)
		sheet_bold_map[sheet_name] = payload.get("bold_list") or []

		ws = wb.create_sheet(title=sheet_name)

		columns = payload.get("columns") or []
		data_rows = payload.get("data") or []

		for col in columns:
			if col.get("fieldtype") in ["Currency", "Float", "Int", "Percent"]:
				right_cell.append(col.get("label"))

		column_widths = payload.get("column_widths")
		for i, column_width in enumerate(column_widths):
			if column_width:
				ws.column_dimensions[get_column_letter(i + 1)].width = column_width

		for row in data_rows:
			ws.append(row)

	bio = BytesIO()
	wb.save(bio)
	bio.seek(0)

	from erpnext.accounts.report.balance_sheet_v2.balance_sheet_v2 import add_formulas

	return add_formulas(
		report_name,
		bio,
		column_widths=column_widths,
		formula=formula,
		bold_list=sheet_bold_map,
		right_cell=right_cell
	)


@frappe.whitelist()
def get_export_with_cost_centers_url(filters=None, formula=True):
	if isinstance(filters, str):
		try:
			filters = frappe.parse_json(filters)
		except Exception:
			filters = {}

	payload = quote(json.dumps(filters or {}))
	url = (
		"/api/method/erpnext.accounts.report.profit_and_loss_malaysia.profit_and_loss_malaysia.export_with_cost_centers"
		f"?filters={payload}&report_name=Profit%20and%20Loss%20Malaysia&formula={cint(formula)}"
	)
	return {"url": url}
