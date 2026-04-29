# Copyright (c) 2015, Frappe Technologies Pvt. Ltd. and Contributors
# License: GNU General Public License v3. See license.txt


import frappe
import re
import json
from urllib.parse import quote
from io import BytesIO
try:
	import openpyxl
except Exception:
	openpyxl = None
from frappe import _
from frappe.utils import flt, now_datetime, get_datetime, now, cint

from erpnext.accounts.report.financial_statements import (
	get_columns,
	get_data,
	get_filtered_list_for_consolidated_report,
	get_period_list,
)
from erpnext.accounts.utils import remove_account_number
from erpnext.accounts.report.utils import convert_wrap_report_data
from erpnext.gp_erp.report.budget_variance_greenphyto.budget_variance_greenphyto import get_budget_account, get_budget_data

def execute(filters=None):
	filters = frappe._dict(filters)
	period_list = get_period_list(
		filters.from_fiscal_year,
		filters.to_fiscal_year,
		filters.period_start_date,
		filters.period_end_date,
		filters.filter_based_on,
		filters.periodicity,
		company=filters.company,
		month=filters.month,
		to_month=filters.to_month,
	)

	filter_zero_value = 1
	accounts_list = []
	# if filters.get("show_budget_amount"):
	# 	filter_zero_value = 0
	# 	if cint(filters.hide_zero_balance):
	# 		accounts_list = get_budget_account(filters.get("cost_center"), filters.get("company"))


	income = get_data(
		filters.company,
		"Income",
		"Credit",
		period_list,
		filters=filters,
		accumulated_values=filters.accumulated_values,
		ignore_closing_entries=True,
		ignore_accumulated_values_for_fy=True,
		filter_zero_value=filter_zero_value,
		accounts_to_show=accounts_list
	)

	expense = get_data(
		filters.company,
		"Expense",
		"Debit",
		period_list,
		filters=filters,
		accumulated_values=filters.accumulated_values,
		ignore_closing_entries=True,
		ignore_accumulated_values_for_fy=True,
		filter_zero_value=filter_zero_value,
		accounts_to_show=accounts_list
	)

	# Fetch budget data and calculate YTD (Year-to-Date) BEFORE extending to data
	budget_map = {}
	# ALWAYS fetch monthly budget data (even for Yearly periodicity)
	# This allows Budget YTD to calculate from raw monthly values
	budget_map = get_budget_data(filters)
	
	# Get current date for YTD limit
	from frappe.utils import today, getdate, add_months
	current_date = getdate(today())
	
	# Add budget to income rows
	if income:
		add_budget_to_rows(income, budget_map, period_list, current_date, filters)
	
	# Add budget to expense rows
	if expense:
		add_budget_to_rows(expense, budget_map, period_list, current_date, filters)
	
	# Calculate net profit/loss AFTER budget is added
	net_profit_loss = get_net_profit_loss(
		income, expense, period_list, filters.company, filters.presentation_currency
	)
	
	data = []
	data.extend(income or [])
	data.extend(expense or [])
	if net_profit_loss:
		data.append(net_profit_loss)

	new_data = []
	if filters.show_number_group:
		new_data = data
	else:
		for d in data:
			if d.get("is_group"):
				d['account_name'] = remove_account_number(d['account_name'])
				if frappe.flags.in_export:
					d['account'] = d['account_name']
					if not d.get('parent_account'):
						for key, val in d.items():
							if key not in ['account', 'account_name']:
								d[key] = None

			new_data.append(d)

	columns = get_columns(
		filters.periodicity, period_list, filters.accumulated_values, filters.company, 
		cost_center_all_show=filters.get("show_all_cost_centers", 0), filters=filters
	)

	chart = get_chart_data(filters, columns, income, expense, net_profit_loss)

	currency = filters.presentation_currency or frappe.get_cached_value(
		"Company", filters.company, "default_currency"
	)
	report_summary = get_report_summary(
		period_list, filters.periodicity, income, expense, net_profit_loss, currency, filters
	)

	# if frappe.flags.in_export:
	# 	convert_wrap_report_data(columns, data, precision=2)

	return columns, new_data, None, chart, report_summary

def get_report_summary(
	period_list, periodicity, income, expense, net_profit_loss, currency, filters, consolidated=False
):
	net_income, net_expense, net_profit = 0.0, 0.0, 0.0

	# from consolidated financial statement
	if filters.get("accumulated_in_group_company"):
		period_list = get_filtered_list_for_consolidated_report(filters, period_list)

	for period in period_list:
		key = period if consolidated else period.key
		if filters.get("accumulated_values"):
			if income:
				net_income = income[-2].get(key)
			if expense:
				net_expense = expense[-2].get(key)
			if net_profit_loss:
				net_profit = net_profit_loss.get(key)
		else:
			if income:
				net_income += income[-2].get(key)
			if expense:
				net_expense += expense[-2].get(key)
			if net_profit_loss:
				net_profit += net_profit_loss.get(key)

	if len(period_list) == 1 and periodicity == "Yearly":
		profit_label = _("Profit This Year")
		income_label = _("Total Income This Year")
		expense_label = _("Total Expense This Year")
	else:
		profit_label = _("Net Profit")
		income_label = _("Total Income")
		expense_label = _("Total Expense")

	return [
		{"value": net_income, "label": income_label, "datatype": "Currency", "currency": currency},
		{"type": "separator", "value": "-"},
		{"value": net_expense, "label": expense_label, "datatype": "Currency", "currency": currency},
		{"type": "separator", "value": "=", "color": "blue"},
		{
			"value": net_profit,
			"indicator": "Green" if net_profit > 0 else "Red",
			"label": profit_label,
			"datatype": "Currency",
			"currency": currency,
		},
	]


def get_net_profit_loss(income, expense, period_list, company, currency=None, consolidated=False):
	total = 0
	net_profit_loss = {
		"account_name": "'" + _("Profit for the year") + "'",
		"account": "'" + _("Profit for the year") + "'",
		"warn_if_negative": True,
		"profit_data":1,
		"currency": currency or frappe.get_cached_value("Company", company, "default_currency"),
	}

	has_value = False

	for period in period_list:
		key = period if consolidated else period.key
		total_income = flt(income[-2][key], 3) if income else 0
		total_expense = flt(expense[-2][key], 3) if expense else 0

		net_profit_loss[key] = total_income - total_expense
		
		# Also calculate budget for net profit/loss (income budget - expense budget)
		budget_key = (period if consolidated else period.key) + "_budget" if not consolidated else None
		if budget_key:
			income_budget = flt(income[-2].get(budget_key, 0), 3) if income else 0
			expense_budget = flt(expense[-2].get(budget_key, 0), 3) if expense else 0
			net_profit_loss[budget_key] = income_budget - expense_budget

		if net_profit_loss[key]:
			has_value = True

		total += flt(net_profit_loss[key])
		net_profit_loss["total"] = total

	if has_value:
		return net_profit_loss


def get_chart_data(filters, columns, income, expense, net_profit_loss):
	# Filter out budget columns from chart data
	period_columns = [d for d in columns[2:] if not d.get("fieldname", "").endswith("_budget")]
	
	labels = [d.get("label") for d in period_columns]

	income_data, expense_data, net_profit = [], [], []

	for p in period_columns:
		if income:
			income_data.append(income[-2].get(p.get("fieldname")))
		if expense:
			expense_data.append(expense[-2].get(p.get("fieldname")))
		if net_profit_loss:
			net_profit.append(net_profit_loss.get(p.get("fieldname")))

	datasets = []
	if income_data:
		datasets.append({"name": _("Income"), "values": income_data})
	if expense_data:
		datasets.append({"name": _("Expense"), "values": expense_data})
	if net_profit:
		datasets.append({"name": _("Net Profit/Loss"), "values": net_profit})

	chart = {"data": {"labels": labels, "datasets": datasets}}

	if not cint(filters.accumulated_values):
		chart["type"] = "bar"
	else:
		chart["type"] = "line"

	chart["fieldtype"] = "Currency"

	return chart

from frappe.desk.query_report import add_title_report, get_filters_data, build_xlsx_data
from openpyxl.utils import get_column_letter
def get_export_cost_center(report_name, filters):
	"""
	Generate grouped Profit & Loss data per non-group Cost Center.

	Returns a dict keyed by Cost Center name with:
	- label: Cost Center display name
	- columns: report columns for this run
	- data: report rows for this Cost Center
	- summary: report summary values
	"""
	base_filters = frappe._dict(filters or {})

	# Collect all non-group cost centers (scoped to company if provided)
	cc_filters = {"is_group": 0}
	if base_filters.get("company"):
		cc_filters["company"] = base_filters.company

	cost_centers = frappe.get_all(
		"Cost Center",
		filters=cc_filters,
		fields=["name", "cost_center_name"],
		order_by="name asc",
	)

	group_data = {}

	# add detail
	export_date = now()
	date_str = " "+get_datetime(export_date).strftime("%-d %B %y %H:%M:%S")
	title_report = add_title_report(report_name) 

	if report_name == "Budget Variance Greenphyto":
		from erpnext.gp_erp.report.budget_variance_greenphyto.budget_variance_greenphyto import execute
	else:
		from erpnext.accounts.report.profit_and_loss_statement.profit_and_loss_statement import execute

	for cc in cost_centers:
		# Prepare per-CC filters without mutating caller filters
		per_cc_filters = frappe._dict(base_filters.copy())
		per_cc_filters["cost_center"] = [cc.name]

		filter_report = get_filters_data(per_cc_filters)
		columns, data, _, _, report_summary = execute(per_cc_filters)
		temp = frappe._dict({
			"columns":columns,
			"result":data
		})
		xlsx_data, column_widths = build_xlsx_data(temp, [] , 1, ignore_visible_idx=1)
		xlsx_data = title_report + filter_report + [["Export date", date_str]] + xlsx_data

		group_data[cc.name] = {
			"label": cc.cost_center_name or cc.name,
			"columns": columns,
			"data": xlsx_data,
			"summary": report_summary,
			"column_widths":column_widths
		}
		# break

	return group_data


def _sanitize_sheet_name(name):
	name = (name or "Sheet").strip()
	# Excel invalid chars: : \ / ? * [ ]
	name = re.sub(r"[:\\/\?\*\[\]]", " ", name)
	if len(name) > 31:
		name = name[:31]
	return name or "Sheet"


@frappe.whitelist()
def export_with_cost_centers(report_name, filters=None, formula=False):
	"""
	Build an XLSX where each Cost Center is a sheet containing its P&L data.

	Returns: { file_url: <saved file url> }
	"""
	if isinstance(filters, str):
		try:
			filters = frappe.parse_json(filters)
		except Exception:
			filters = {}

	formula = cint(formula)
	filters['show_all_cost_centers'] = 0
	group_data  = get_export_cost_center(report_name, filters)

	if not openpyxl:
		frappe.throw("openpyxl is required to export XLSX")

	from frappe.utils.xlsxutils import ILLEGAL_CHARACTERS_RE, handle_html

	wb = openpyxl.Workbook(write_only=True)

	used_names = set()
	first_columns = None

	for cc_name, payload in (group_data or {}).items():
		label = payload.get("label") or cc_name
		sheet_name = _sanitize_sheet_name(label)

		# Ensure unique sheet names
		base = sheet_name
		idx = 1
		while sheet_name in used_names:
			suffix = f" {idx}"
			sheet_name = _sanitize_sheet_name((base[: (31 - len(suffix))]).rstrip() + suffix)
			idx += 1
		used_names.add(sheet_name)

		ws = wb.create_sheet(title=sheet_name)

		columns = payload.get("columns") or []
		if first_columns is None:
			first_columns = columns
		data_rows = payload.get("data") or []

		headers = [c.get("label") for c in columns]
		fields = [c.get("fieldname") for c in columns]
		ws.append(headers)

		# Set column width if provided
		column_widths = payload.get("column_widths")
		for i, column_width in enumerate(column_widths):
			if column_width:
				ws.column_dimensions[get_column_letter(i + 1)].width = column_width

		for row in data_rows:
			out = []
			if type(row) != list:
				for f in fields:
					val = (row or {}).get(f)
					if isinstance(val, str):
						try:
							val = handle_html(val)
							if isinstance(val, str):
								val = re.sub(ILLEGAL_CHARACTERS_RE, "", val)
						except Exception:
							pass
					out.append(val)
			else:
				out += row

			ws.append(out)

	bio = BytesIO()
	wb.save(bio)
	bio.seek(0)

	# Integrate with add_formulas to add formulas on all sheets
	try:
		from erpnext.accounts.report.balance_sheet_v2.balance_sheet_v2 import add_formulas
		# Build column_widths list from the first sheet's columns if available
		column_widths = None
		if first_columns:
			column_widths = []
			for col in first_columns:
				w = col.get("width")
				# Convert typical pixel widths to character widths approximation if needed
				if isinstance(w, (int, float)):
					# heuristic: 1 char ~ 8 px; clamp sensible range
					width_chars = max(6, min(80, int(float(w) / 8)))
				else:
					# fall back to label length approx
					label = col.get("label") or ""
					width_chars = max(8, min(50, len(str(label)) + 5))
				column_widths.append(width_chars)

		return add_formulas(report_name, bio, column_widths=column_widths, formula=formula)

	except Exception:
		# Fallback: plain export if add_formulas unavailable
		now = now_datetime()
		date_str_title = now.strftime("%y%m%d_%H%M%S")
		frappe.response["filename"] = f"Profit and Loss by Cost Center_{date_str_title}.xlsx"
		frappe.response["filecontent"] = bio.getvalue()
		frappe.response["type"] = "binary"


@frappe.whitelist()
def get_export_with_cost_centers_url(filters=None, formula=True):
	"""
	Helper to generate a URL for binary export so that the client
	can use frappe.call() first, then window.open() the returned URL.
	"""
	if isinstance(filters, str):
		try:
			filters = frappe.parse_json(filters)
		except Exception:
			filters = {}

	payload = quote(json.dumps(filters or {}))
	url = (
		"/api/method/erpnext.accounts.report.profit_and_loss_statement.profit_and_loss_statement.export_with_cost_centers"
		f"?filters={payload}&report_name=Profit%20and%20Loss%20Statement&formula={cint(formula)}"
	)
	return {"url": url}

def add_budget_to_rows(rows, budget_map, period_list, current_date, filters):
	"""Add budget amounts to all rows including totals
	
	Args:
		rows: List of account rows (income or expense)
		budget_map: Dict mapping account -> period -> budget_amount
		period_list: List of periods to add budget for
		current_date: Current date for YTD limit
		filters: Report filters
	"""
	from frappe.utils import getdate, add_months, flt
	
	# First pass: Add budget to leaf accounts (non-group, non-total)
	for row in rows:
		# Skip total rows (they'll be calculated later)
		account_name = row.get("account_name", "").lower()
		is_total_row = "total" in account_name and ("income" in account_name or "expense" in account_name)
		
		if is_total_row:
			continue
		
		# Skip group accounts in first pass
		if row.get("is_group"):
			continue
			
		account = row.get("account_origin") or row.get("account")
		if not account:
			continue
		
		# Add budget for each period
		for period in period_list:
			budget_key = period.key + "_budget"
			budget_value = 0
			
			period_to_date = getdate(period.to_date) if period.to_date else None
			
			if filters.periodicity == "Yearly":
				# For yearly budget, sum all 12 months from budget_map
				# budget_map is now always monthly: account -> month_number -> budget_amount
				if account in budget_map:
					for month_num in range(1, 13):
						budget_value += flt(budget_map.get(account, {}).get(month_num, 0))
			else:
				# Monthly budget
				if account in budget_map and period_to_date:
					if filters.accumulated_values:
						# Accumulated mode: Calculate YTD
						fiscal_year_start = getdate(period.year_start_date) if period.year_start_date else None
						
						if fiscal_year_start:
							ytd_budget = 0
							current_month_date = fiscal_year_start
							
							# Iterate month by month from fiscal year start to period end
							while current_month_date <= period_to_date:
								month_num = current_month_date.month
								ytd_budget += budget_map.get(account, {}).get(month_num, 0)
								
								# Move to next month
								current_month_date = add_months(current_month_date, 1)
								
								# Safety check
								if current_month_date.year > period_to_date.year or \
								   (current_month_date.year == period_to_date.year and current_month_date.month > period_to_date.month):
									break
							
							budget_value = ytd_budget
					else:
						# Non-accumulated mode: Show budget for this month only
						month_num = period_to_date.month
						budget_value = budget_map.get(account, {}).get(month_num, 0)
			
			row[budget_key] = budget_value
	
	# Second pass: Calculate budget for group accounts (bottom-up)
	for row in reversed(rows):
		account_name = row.get("account_name", "").lower()
		is_total_row = "total" in account_name and ("income" in account_name or "expense" in account_name)
		
		if is_total_row:
			continue
		
		if row.get("is_group"):
			account = row.get("account_origin") or row.get("account")
			if not account:
				continue
			
			# Initialize budget totals for this group
			for period in period_list:
				budget_key = period.key + "_budget"
				row[budget_key] = 0
			
			# Sum budget from child accounts
			for child_row in rows:
				child_account_name = child_row.get("account_name", "").lower()
				child_is_total = "total" in child_account_name and ("income" in child_account_name or "expense" in child_account_name)
				
				if child_is_total:
					continue
				
				child_parent = child_row.get("parent_account")
				if child_parent == account:
					for period in period_list:
						budget_key = period.key + "_budget"
						child_budget = flt(child_row.get(budget_key, 0))
						row[budget_key] = flt(row.get(budget_key, 0)) + child_budget
	
	# Third pass: Calculate budget for total row (it's the second to last row, -2)
	if len(rows) >= 2:
		total_row = rows[-2]
		account_name = total_row.get("account_name", "").lower()
		is_total_row = "total" in account_name and ("income" in account_name or "expense" in account_name)
		
		if is_total_row:
			# Initialize all budget periods
			for period in period_list:
				budget_key = period.key + "_budget"
				total_row[budget_key] = 0
			
			# Sum budget from all top-level accounts (accounts with no parent or root parent)
			for row in rows[:-2]:  # Exclude last 2 rows (total and separator)
				row_account_name = row.get("account_name", "").lower()
				row_is_total = "total" in row_account_name and ("income" in row_account_name or "expense" in row_account_name)
				
				if row_is_total:
					continue
				
				parent = row.get("parent_account")
				# Only sum top-level accounts (no parent, or parent is root like "Income" or "Expense")
				if not parent or parent in ["Income", "Expenses", "Expense"]:
					for period in period_list:
						budget_key = period.key + "_budget"
						row_budget = flt(row.get(budget_key, 0))
						total_row[budget_key] = flt(total_row.get(budget_key, 0)) + row_budget