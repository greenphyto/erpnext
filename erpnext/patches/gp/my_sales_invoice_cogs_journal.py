import os

import frappe
import openpyxl
from frappe.utils import flt


COMPANY = "Greenphyto Tech Sdn Bhd"
POSTING_DATE = "2026-08-31"
EXCEL_PATH = os.path.expanduser("~/MY_SALES_INVOICE_COGS.xlsx")
COGS_ACCOUNT = "500000 - COGS - Direct - GTSB"
WASTE_ACCOUNT = "621700 - Marketing Inventory Utilisation - GTSB"


# bench --site erp-prod execute erpnext.patches.gp.my_sales_invoice_cogs_journal.execute
def execute():
    total_cogs = get_total_cogs()
    total_cogs = flt(total_cogs, 2)
    if total_cogs <= 0:
        frappe.msgprint("No positive COGS amount found")
        return

    cost_center = frappe.db.get_value("Company", COMPANY, "cost_center")
    je = frappe.new_doc("Journal Entry")
    je.voucher_type = "Journal Entry"
    je.posting_date = POSTING_DATE
    je.company = COMPANY
    je.user_remark = f"Transfer missing Sales Invoice COGS from Waste as at {POSTING_DATE}"
    je.append("accounts", {
        "account": COGS_ACCOUNT,
        "debit": total_cogs,
        "debit_in_account_currency": total_cogs,
        "cost_center": cost_center,
        "user_remark": "Sales Invoice COGS from MY_SALES_INVOICE_COGS.xlsx",
    })
    je.append("accounts", {
        "account": WASTE_ACCOUNT,
        "credit": total_cogs,
        "credit_in_account_currency": total_cogs,
        "cost_center": cost_center,
        "user_remark": "Transfer out of Waste inventory utilisation",
    })
    je.insert()
    frappe.db.commit()
    frappe.msgprint(f"Journal Entry draft: {je.name}\nAmount: {total_cogs:.2f}")


def get_total_cogs():
    workbook = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    worksheet = workbook["Item Breakdown"]
    headers = [cell.value for cell in next(worksheet.iter_rows())]
    cogs_column = headers.index("COGS Amount")
    return sum(flt(row[cogs_column], 2) for row in worksheet.iter_rows(min_row=2, values_only=True))


if __name__ == "__main__":
    execute()
