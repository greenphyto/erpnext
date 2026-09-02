import os

import frappe
import openpyxl
from openpyxl.styles import Font


COMPANY = "Greenphyto Tech Sdn Bhd"
KOKUBU = "Kokubu warehouse - GTSB"
CONSIGNMENT = "Consignment - VILLAGE GROCER - GTSB"
WAREHOUSES = [KOKUBU, CONSIGNMENT]
FROM_DATE = "2026-01-01"
TO_DATE = "2026-08-31"
OUTPUT_PATH = os.path.expanduser("~/MY_SALES_INVOICE_COGS.xlsx")
DETAIL_HEADERS = [
    "No.", "Sales Invoice", "Type", "Posting Date", "Item Code", "Item Name",
    "Stock Qty", "Selling Rate", "Selling Amount", "COGS Rate", "COGS Amount",
]

# bench --site erp-prod execute erpnext.patches.gp.my_sales_invoice_cogs.execute
def execute():
    rates = get_purchase_receipt_rates()
    detail_rows = get_sales_invoice_rows()
    workbook = openpyxl.Workbook()
    detail = workbook.active
    detail.title = "Item Breakdown"
    append_rows(detail, DETAIL_HEADERS, detail_rows)
    total_row = len(detail_rows) + 3
    detail.cell(total_row, 1, "Total")
    detail.cell(total_row, 7, f"=SUM(G2:G{total_row - 2})")
    detail.cell(total_row, 9, f"=SUM(I2:I{total_row - 2})")
    detail.cell(total_row, 11, f"=SUM(K2:K{total_row - 2})")
    for cell in detail[total_row]:
        cell.font = Font(bold=True)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells) + 2
            sheet.column_dimensions[column_cells[0].column_letter].width = min(width, 45)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0" if cell.column == 1 else "0.00"
    workbook.save(OUTPUT_PATH)
    frappe.msgprint(f"Output written to {OUTPUT_PATH}\nRows: {len(detail_rows)}\nTotal COGS: {sum(row[10] for row in detail_rows):.2f}")


def get_purchase_receipt_rates():
    rows = frappe.db.sql(
        """
        SELECT sle.warehouse, sle.item_code,
               SUM(sle.actual_qty) AS qty,
               SUM(sle.stock_value_difference) AS value
        FROM `tabStock Ledger Entry` sle
        INNER JOIN `tabPurchase Receipt` pr ON pr.name = sle.voucher_no
        WHERE sle.company = %s
          AND sle.warehouse IN %s
          AND sle.posting_date BETWEEN %s AND %s
          AND sle.voucher_type = 'Purchase Receipt'
          AND sle.actual_qty > 0
          AND sle.is_cancelled = 0
          AND pr.docstatus = 1
        GROUP BY sle.warehouse, sle.item_code
        HAVING SUM(sle.actual_qty) > 0
        """,
        (COMPANY, (KOKUBU, CONSIGNMENT), FROM_DATE, TO_DATE),
        as_dict=True,
    )
    rates = {KOKUBU: {}, CONSIGNMENT: {}}
    for row in rows:
        rates[row.warehouse][row.item_code] = flt(row.value / row.qty)
    rates[CONSIGNMENT] = rates[KOKUBU].copy()
    return rates


def get_sales_invoice_rows():
    rows = frappe.db.sql(
        """
        SELECT si.name, si.posting_date, si.is_return, sii.warehouse, sii.item_code,
               sii.item_name, sii.stock_qty, sii.amount
        FROM `tabSales Invoice` si
        INNER JOIN `tabSales Invoice Item` sii ON sii.parent = si.name
        WHERE si.company = %s
          AND si.posting_date BETWEEN %s AND %s
          AND si.docstatus = 1
          AND (si.update_stock = 0 OR si.is_return = 1)
          AND sii.item_code IS NOT NULL
          AND sii.item_code != ''
          AND sii.item_code != 'Debit Note'
        ORDER BY si.is_return, si.posting_date, si.name, sii.warehouse, sii.item_code
        """,
        (COMPANY, FROM_DATE, TO_DATE),
        as_dict=True,
    )
    rates = get_purchase_receipt_rates()
    result = []
    for row in rows:
        qty = flt(row.stock_qty)
        if row.is_return:
            qty = -abs(qty)
        rate = rates.get(KOKUBU, {}).get(row.item_code, 0)
        selling_rate = flt(row.amount / qty) if qty else 0
        result.append([
            int(len(result) + 1), row.name, "Return" if row.is_return else "Invoice",
            str(row.posting_date), row.item_code, row.item_name, qty,
            selling_rate, flt(row.amount), rate, flt(qty * rate),
        ])
    return result


def append_rows(sheet, headers, rows):
    sheet.append(headers)
    for row in rows:
        sheet.append(row)


def flt(value):
    return frappe.utils.flt(value, 2)


if __name__ == "__main__":
    execute()
