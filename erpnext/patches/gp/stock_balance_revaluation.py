import frappe
from frappe.utils import flt
from frappe.utils.xlsxutils import make_xlsx
import openpyxl
from io import BytesIO
import os


def execute():
    site = frappe.local.site
    warehouse = "Finished Goods - GPL"
    item_group = "Products"
    cutoff_date = "2026-01-01"
    end_date = "2026-08-15"

    item_name_map = get_item_name_map(item_group)
    account_map = get_account_map()
    gl_balance_map = get_gl_balance_map(end_date)
    opening_map = get_opening_balance(warehouse, item_group, cutoff_date)

    manufacturing = get_manufacturing(warehouse, item_group, cutoff_date, end_date)
    dn_return = get_dn_return(warehouse, item_group, cutoff_date, end_date)
    delivery_note = get_delivery_note(warehouse, item_group, cutoff_date, end_date)
    scrap = get_scrap(warehouse, item_group, cutoff_date, end_date)
    reconciliation = get_reconciliation(warehouse, item_group, cutoff_date, end_date)

    items = set()
    for row in manufacturing:
        items.add(row.item_code)
    for row in dn_return:
        items.add(row.item_code)
    for row in delivery_note:
        items.add(row.item_code)
    for row in scrap:
        items.add(row.item_code)
    for row in reconciliation:
        items.add(row.item_code)

    mfg_by_item = group_by_item(manufacturing)
    dnr_by_item = group_by_item(dn_return)
    dn_by_item = group_by_item(delivery_note)
    scrap_by_item = group_by_item(scrap)
    recon_by_item = group_by_item(reconciliation)

    summary_rows = []
    mfg_detail_rows = []
    dnr_detail_rows = []
    dn_detail_rows = []
    scrap_detail_rows = []
    recon_detail_rows = []

    for item_code in sorted(items):
        mfg_entries = mfg_by_item.get(item_code, [])
        dnr_entries = dnr_by_item.get(item_code, [])
        dn_entries = dn_by_item.get(item_code, [])
        scrap_entries = scrap_by_item.get(item_code, [])
        recon_entries = recon_by_item.get(item_code, [])

        mfg_qty = sum(flt(r.actual_qty) for r in mfg_entries)
        mfg_value = sum(flt(r.stock_value_difference) for r in mfg_entries)

        avg_rate = flt(mfg_value / mfg_qty, 10) if mfg_qty else 0

        dnr_qty = sum(flt(r.actual_qty) for r in dnr_entries if flt(r.actual_qty) > 0)
        dnr_qty += sum(flt(r.actual_qty) for r in dn_entries if flt(r.actual_qty) > 0)
        dnr_value = flt(dnr_qty * avg_rate)

        dn_qty = sum(abs(flt(r.actual_qty)) for r in dn_entries if flt(r.actual_qty) < 0)
        dn_value = flt(dn_qty * avg_rate)

        scrap_qty = sum(abs(flt(r.actual_qty)) for r in scrap_entries)
        scrap_value = flt(scrap_qty * avg_rate)

        recon_qty = sum(flt(r.actual_qty) for r in recon_entries)
        recon_value = sum(flt(r.stock_value_difference) for r in recon_entries)

        opening = opening_map.get(item_code, {})
        opening_qty = flt(opening.get("qty", 0))
        opening_value = flt(opening.get("value", 0))

        total_incoming_qty = mfg_qty + dnr_qty
        total_incoming_value = mfg_value + dnr_value
        total_outgoing_qty = dn_qty + scrap_qty
        total_outgoing_value = flt(total_outgoing_qty * avg_rate)

        current = get_current_stock(item_code, warehouse, end_date)
        current_qty = flt(current.get("actual_qty", 0))

        account = account_map.get(item_code, "")
        current_account_balance = flt(gl_balance_map.get(account, 0))

        item_name = item_name_map.get(item_code, item_code)

        summary_rows.append([
            item_code,
            item_name,
            account,
            flt(avg_rate, 2),
            flt(opening_qty, 2),
            flt(opening_value, 2),
            flt(mfg_qty, 2),
            flt(mfg_value, 2),
            flt(dnr_qty, 2),
            None,
            flt(dn_qty, 2),
            None,
            flt(scrap_qty, 2),
            None,
            flt(recon_qty, 2),
            None,
            flt(total_outgoing_value, 2),
            None,
            None,
            flt(current_qty, 2),
            flt(current_account_balance, 2),
            None,
        ])

        for r in mfg_entries:
            mfg_detail_rows.append([
                str(r.posting_date),
                r.voucher_no,
                r.item_code,
                item_name,
                flt(r.actual_qty, 2),
                flt(r.valuation_rate, 2),
                flt(r.stock_value_difference, 2),
            ])

        for r in dnr_entries:
            cur_rate = flt(r.doc_rate, 2)
            cur_amount = flt(flt(r.actual_qty) * cur_rate, 2)
            correct_amount = flt(flt(r.actual_qty) * avg_rate, 2)
            diff = flt(cur_amount - correct_amount, 2)
            dnr_detail_rows.append([
                str(r.posting_date),
                r.voucher_no,
                r.item_code,
                item_name,
                flt(r.actual_qty, 2),
                cur_rate,
                cur_amount,
                flt(avg_rate, 2),
                correct_amount,
                diff,
            ])

        for r in dn_entries:
            cur_rate = flt(r.doc_rate, 2)
            cur_amount = flt(abs(flt(r.actual_qty)) * cur_rate, 2)
            correct_amount = flt(abs(flt(r.actual_qty)) * avg_rate, 2)
            diff = flt(cur_amount - correct_amount, 2)
            dn_detail_rows.append([
                str(r.posting_date),
                r.voucher_no,
                r.item_code,
                item_name,
                flt(abs(flt(r.actual_qty)), 2),
                cur_rate,
                cur_amount,
                flt(avg_rate, 2),
                correct_amount,
                diff,
            ])

        for r in scrap_entries:
            cur_rate = flt(r.doc_rate, 2)
            cur_amount = flt(r.doc_amount, 2) if r.doc_amount else flt(abs(flt(r.actual_qty)) * cur_rate, 2)
            correct_amount = flt(abs(flt(r.actual_qty)) * avg_rate, 2)
            diff = flt(cur_amount - correct_amount, 2)
            scrap_detail_rows.append([
                str(r.posting_date),
                r.voucher_no,
                r.item_code,
                item_name,
                flt(abs(flt(r.actual_qty)), 2),
                cur_rate,
                cur_amount,
                flt(avg_rate, 2),
                correct_amount,
                diff,
            ])

        for r in recon_entries:
            recon_detail_rows.append([
                str(r.posting_date),
                r.voucher_no,
                r.item_code,
                item_name,
                flt(r.actual_qty, 2),
                flt(r.valuation_rate, 2),
                flt(r.stock_value_difference, 2),
            ])

    mfg_detail_rows.sort(key=lambda x: (x[0], x[1], x[2]))
    dnr_detail_rows.sort(key=lambda x: (x[0], x[1], x[2]))
    dn_detail_rows.sort(key=lambda x: (x[0], x[1], x[2]))
    scrap_detail_rows.sort(key=lambda x: (x[0], x[1], x[2]))
    recon_detail_rows.sort(key=lambda x: (x[0], x[1], x[2]))

    summary_header = [
        "Item Code", "Item Name", "BS Account", "Avg Rate (Mfg)",
        "Opening Qty", "Opening Value",
        "Mfg Qty", "Mfg Value",
        "DN Return Qty", "DN Return Value",
        "DN Out Qty", "DN Out Value",
        "Scrap Qty", "Scrap Value",
        "Recon Qty", "Recon Value",
        "Total Outgoing Value",
        "Balance Qty", "Balance Value",
        "Current Qty (Bin)", "Current Account Balance", "Diff",
    ]
    detail_header = [
        "Posting Date", "Voucher No", "Item Code", "Item Name",
        "Qty", "Rate", "Amount",
    ]
    detail_header_compare = [
        "Posting Date", "Voucher No", "Item Code", "Item Name",
        "Qty", "Cur Rate", "Cur Amount", "Correct Rate", "Correct Amount", "Diff",
    ]

    summary_data = [summary_header] + summary_rows
    mfg_data = [detail_header] + mfg_detail_rows
    dnr_data = [detail_header_compare] + dnr_detail_rows
    dn_data = [detail_header_compare] + dn_detail_rows
    scrap_data = [detail_header_compare] + scrap_detail_rows
    recon_data = [detail_header] + recon_detail_rows

    wb = openpyxl.Workbook(write_only=False)
    if wb.active and wb.active.title == "Sheet":
        wb.remove(wb.active)

    make_xlsx(summary_data, "Summary", wb=wb)
    make_xlsx(mfg_data, "Manufacturing", wb=wb)
    make_xlsx(dnr_data, "DN Return", wb=wb)
    make_xlsx(dn_data, "Delivery Note", wb=wb)
    make_xlsx(scrap_data, "Scrap Materials", wb=wb)
    make_xlsx(recon_data, "Reconciliation", wb=wb)

    ws = wb["Summary"]
    for row_idx in range(2, len(summary_rows) + 2):
        ws.cell(row=row_idx, column=10, value=f"=I{row_idx}*D{row_idx}")
        ws.cell(row=row_idx, column=12, value=f"=K{row_idx}*D{row_idx}")
        ws.cell(row=row_idx, column=14, value=f"=M{row_idx}*D{row_idx}")
        ws.cell(row=row_idx, column=16, value=f"=O{row_idx}*D{row_idx}")
        ws.cell(row=row_idx, column=18, value=f"=E{row_idx}+G{row_idx}+I{row_idx}-K{row_idx}-M{row_idx}+O{row_idx}")
        ws.cell(row=row_idx, column=19, value=f"=R{row_idx}*D{row_idx}")
        ws.cell(row=row_idx, column=22, value=f"=S{row_idx}-U{row_idx}")

    output_path = os.path.expanduser("~/STOCK_BALANCE_REVALUATION.xlsx")
    wb.save(output_path)
    frappe.msgprint(f"Output written to {output_path}\nItems: {len(summary_rows)}")


def get_item_name_map(item_group):
    items = frappe.db.sql("""
        SELECT name, item_name FROM `tabItem` WHERE item_group = %s
    """, (item_group,), as_dict=True)
    return {i.name: i.item_name for i in items}


def get_account_map():
    results = frappe.db.sql("""
        SELECT code, account_code
        FROM `tabPart Number Details`
        WHERE account_code LIKE '%%GPL%%'
        AND code LIKE 'PR-%%'
        AND account_code IS NOT NULL AND account_code != ''
    """, as_dict=True)
    return {r.code: r.account_code for r in results}


def get_gl_balance_map(end_date):
    results = frappe.db.sql("""
        SELECT account, SUM(debit) - SUM(credit) as balance
        FROM `tabGL Entry`
        WHERE account LIKE '121%%GPL'
          AND posting_date <= %s
          AND is_cancelled = 0
        GROUP BY account
    """, (end_date,), as_dict=True)
    return {r.account: r.balance for r in results}


def get_opening_balance(warehouse, item_group, cutoff_date):
    results = frappe.db.sql("""
        SELECT sle.item_code, sle.qty_after_transaction as qty, sle.stock_value as value
        FROM `tabStock Ledger Entry` sle
        JOIN `tabItem` i ON i.name = sle.item_code
        WHERE sle.warehouse = %s
          AND i.item_group = %s
          AND sle.posting_date < %s
          AND sle.is_cancelled = 0
          AND sle.name = (
              SELECT s2.name FROM `tabStock Ledger Entry` s2
              WHERE s2.item_code = sle.item_code
                AND s2.warehouse = sle.warehouse
                AND s2.posting_date < %s
                AND s2.is_cancelled = 0
              ORDER BY s2.posting_date DESC, s2.posting_time DESC, s2.creation DESC
              LIMIT 1
          )
    """, (warehouse, item_group, cutoff_date, cutoff_date), as_dict=True)
    return {r.item_code: {"qty": r.qty, "value": r.value} for r in results}


def get_manufacturing(warehouse, item_group, cutoff_date, end_date):
    return frappe.db.sql("""
        SELECT
            sle.posting_date, sle.voucher_no, sle.item_code,
            sle.actual_qty, sle.valuation_rate, sle.stock_value_difference
        FROM `tabStock Ledger Entry` sle
        JOIN `tabItem` i ON i.name = sle.item_code
        JOIN `tabStock Entry` se ON se.name = sle.voucher_no
        WHERE sle.warehouse = %s
          AND i.item_group = %s
          AND sle.posting_date >= %s
          AND sle.posting_date <= %s
          AND sle.is_cancelled = 0
          AND sle.voucher_type = 'Stock Entry'
          AND sle.actual_qty > 0
          AND se.purpose = 'Manufacture'
        ORDER BY sle.posting_date, sle.posting_time, sle.creation
    """, (warehouse, item_group, cutoff_date, end_date), as_dict=True)


def get_dn_return(warehouse, item_group, cutoff_date, end_date):
    return frappe.db.sql("""
        SELECT
            sle.posting_date, sle.voucher_no, sle.item_code,
            sle.actual_qty, sle.valuation_rate, sle.stock_value_difference,
            sle.voucher_detail_no,
            dni.rate as doc_rate
        FROM `tabStock Ledger Entry` sle
        JOIN `tabItem` i ON i.name = sle.item_code
        JOIN `tabDelivery Note` dn ON dn.name = sle.voucher_no
        LEFT JOIN `tabDelivery Note Item` dni ON dni.name = sle.voucher_detail_no
        WHERE sle.warehouse = %s
          AND i.item_group = %s
          AND sle.posting_date >= %s
          AND sle.posting_date <= %s
          AND sle.is_cancelled = 0
          AND sle.voucher_type = 'Delivery Note'
          AND dn.is_return = 1
        ORDER BY sle.posting_date, sle.posting_time, sle.creation
    """, (warehouse, item_group, cutoff_date, end_date), as_dict=True)


def get_delivery_note(warehouse, item_group, cutoff_date, end_date):
    return frappe.db.sql("""
        SELECT
            sle.posting_date, sle.voucher_no, sle.item_code,
            sle.actual_qty, sle.valuation_rate, sle.stock_value_difference,
            sle.voucher_detail_no,
            dni.rate as doc_rate
        FROM `tabStock Ledger Entry` sle
        JOIN `tabItem` i ON i.name = sle.item_code
        JOIN `tabDelivery Note` dn ON dn.name = sle.voucher_no
        LEFT JOIN `tabDelivery Note Item` dni ON dni.name = sle.voucher_detail_no
        WHERE sle.warehouse = %s
          AND i.item_group = %s
          AND sle.posting_date >= %s
          AND sle.posting_date <= %s
          AND sle.is_cancelled = 0
          AND sle.voucher_type = 'Delivery Note'
          AND dn.is_return = 0
        ORDER BY sle.posting_date, sle.posting_time, sle.creation
    """, (warehouse, item_group, cutoff_date, end_date), as_dict=True)


def get_scrap(warehouse, item_group, cutoff_date, end_date):
    return frappe.db.sql("""
        SELECT
            sle.posting_date, sle.voucher_no, sle.item_code,
            sle.actual_qty, sle.valuation_rate, sle.stock_value_difference,
            sle.voucher_detail_no,
            sed.basic_rate as doc_rate, sed.basic_amount as doc_amount
        FROM `tabStock Ledger Entry` sle
        JOIN `tabItem` i ON i.name = sle.item_code
        JOIN `tabStock Entry` se ON se.name = sle.voucher_no
        LEFT JOIN `tabStock Entry Detail` sed ON sed.name = sle.voucher_detail_no
        WHERE sle.warehouse = %s
          AND i.item_group = %s
          AND sle.posting_date >= %s
          AND sle.posting_date <= %s
          AND sle.is_cancelled = 0
          AND sle.voucher_type = 'Stock Entry'
          AND sle.actual_qty < 0
          AND (se.stock_entry_type_view = 'Scrap Materials' OR se.purpose = 'Material Issue')
        ORDER BY sle.posting_date, sle.posting_time, sle.creation
    """, (warehouse, item_group, cutoff_date, end_date), as_dict=True)


def get_current_stock(item_code, warehouse, end_date):
    results = frappe.db.sql("""
        SELECT sle.qty_after_transaction as actual_qty, sle.stock_value
        FROM `tabStock Ledger Entry` sle
        WHERE sle.item_code = %s
          AND sle.warehouse = %s
          AND sle.posting_date <= %s
          AND sle.is_cancelled = 0
        ORDER BY sle.posting_date DESC, sle.posting_time DESC, sle.creation DESC
        LIMIT 1
    """, (item_code, warehouse, end_date), as_dict=True)
    if results:
        r = results[0]
        return {"actual_qty": r.actual_qty, "stock_value": r.stock_value}
    return {"actual_qty": 0, "stock_value": 0}


def get_reconciliation(warehouse, item_group, cutoff_date, end_date):
    return frappe.db.sql("""
        SELECT
            sle.posting_date, sle.voucher_no, sle.item_code,
            sle.actual_qty, sle.valuation_rate, sle.stock_value_difference
        FROM `tabStock Ledger Entry` sle
        JOIN `tabItem` i ON i.name = sle.item_code
        WHERE sle.warehouse = %s
          AND i.item_group = %s
          AND sle.posting_date >= %s
          AND sle.posting_date <= %s
          AND sle.is_cancelled = 0
          AND sle.voucher_type = 'Stock Reconciliation'
        ORDER BY sle.posting_date, sle.posting_time, sle.creation
    """, (warehouse, item_group, cutoff_date, end_date), as_dict=True)


def group_by_item(entries):
    result = {}
    for r in entries:
        result.setdefault(r.item_code, []).append(r)
    return result
