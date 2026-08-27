import os

import frappe
import openpyxl


COMPANY = "Greenphyto Pte Ltd"
WAREHOUSE = "Finished Goods - GPL"
ITEM_GROUP = "Products"
FROM_DATE = "2026-01-01"
TO_DATE = "2026-08-15"


STOCK_HEADERS = [
    "Item Code", "Item Name", "Item Group", "Warehouse", "Account", "Acc Code",
    "Balance Qty", "Balance Value", "Opening Qty", "Opening Value", "In Qty", "In Value",
    "Out Qty", "Out Value", "Valuation Rate",
]
MANUFACTURING_HEADERS = ["Item Code", "Item Name", "Manufacturing Qty", "Manufacturing Value", "Avg Rate"]
SUMMARY_HEADERS = [
    "Item Code", "Item Name", "BS Account", "Stock Balance Qty", "Manufacturing Qty",
    "Manufacturing Value", "New Rate", "Calculated Balance", "BS Account Current Balance", "Difference",
]


def execute():
    stock_rows, manufacturing_rows = get_data()
    gl_balances = get_gl_balances()
    workbook = openpyxl.Workbook()
    stock_sheet = workbook.active
    stock_sheet.title = "Stock Balance"
    manufacturing_sheet = workbook.create_sheet("Manufacturing Value")
    summary_sheet = workbook.create_sheet("Summary")

    append_sheet(stock_sheet, STOCK_HEADERS, stock_rows)
    append_sheet(manufacturing_sheet, MANUFACTURING_HEADERS, manufacturing_rows)

    summary_sheet.append(SUMMARY_HEADERS)
    item_codes = sorted(set(stock_rows) | set(manufacturing_rows))
    for item_code in item_codes:
        row = summary_sheet.max_row + 1
        item_name = stock_rows.get(item_code, manufacturing_rows.get(item_code, {})).get("item_name", "")
        bs_account = stock_rows.get(item_code, {}).get("acc_code", "")
        summary_sheet.append([item_code, item_name, bs_account, None, None, None, None, None, gl_balances.get(bs_account, 0), None])
        summary_sheet.cell(row, 4, f'=SUMIF(\'Stock Balance\'!$A:$A,A{row},\'Stock Balance\'!$G:$G)')
        summary_sheet.cell(row, 5, f'=SUMIF(\'Manufacturing Value\'!$A:$A,A{row},\'Manufacturing Value\'!$C:$C)')
        summary_sheet.cell(row, 6, f'=SUMIF(\'Manufacturing Value\'!$A:$A,A{row},\'Manufacturing Value\'!$D:$D)')
        summary_sheet.cell(row, 7, f"=IFERROR(F{row}/E{row},0)")
        summary_sheet.cell(row, 8, f"=D{row}*G{row}")
        summary_sheet.cell(row, 10, f"=H{row}-I{row}")

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                    cell.number_format = "0.00"

    output_path = os.path.expanduser("~/STOCK_BALANCE_REVALUATION_V2.xlsx")
    workbook.save(output_path)
    frappe.msgprint(f"Output written to {output_path}\nItems: {len(item_codes)}")


def get_data():
    entries = frappe.db.sql(
        """
        SELECT sle.item_code, i.item_name, i.item_group, sle.warehouse, sle.batch_no,
            sle.posting_date, sle.posting_time, sle.creation, sle.name, sle.actual_qty,
            sle.qty_after_transaction, sle.stock_value_difference, sle.stock_value,
            sle.valuation_rate,             sle.voucher_type, sle.voucher_no,
            se.purpose AS stock_entry_purpose,
            CASE WHEN sle.voucher_type = 'Stock Reconciliation' THEN sr.purpose ELSE '' END AS reconciliation_purpose,
            pnd.account_code, acc.account_name, acc.account_number
        FROM `tabStock Ledger Entry` sle
        JOIN `tabItem` i ON i.name = sle.item_code
        LEFT JOIN `tabStock Entry` se ON se.name = sle.voucher_no AND sle.voucher_type = 'Stock Entry'
        LEFT JOIN `tabStock Reconciliation` sr ON sr.name = sle.voucher_no

        LEFT JOIN `tabPart Number Details` pnd ON pnd.parent = %s AND pnd.code = sle.item_code
        LEFT JOIN `tabAccount` acc ON acc.name = pnd.account_code
        WHERE sle.company = %s AND sle.warehouse = %s AND i.item_group = %s
          AND sle.posting_date >= '1900-01-01' AND sle.posting_date <= %s
          AND sle.docstatus < 2 AND sle.is_cancelled = 0
        ORDER BY sle.item_code, sle.batch_no, sle.posting_date, sle.posting_time, sle.creation, sle.actual_qty, sle.name
        """,
        (COMPANY, COMPANY, WAREHOUSE, ITEM_GROUP, TO_DATE),
        as_dict=True,
    )
    stock = {}
    manufacturing = {}
    states = {}
    latest = {}
    for entry in entries:
        item = entry.item_code
        batch = entry.batch_no or ""
        key = (item, batch)
        state = states.setdefault(key, {"qty": 0, "value": 0})
        qty = entry.actual_qty or 0
        value = entry.stock_value_difference or 0
        if entry.voucher_type == "Stock Reconciliation" and not entry.batch_no:
            qty_diff = (entry.qty_after_transaction or 0) - sum(s["qty"] for k, s in states.items() if k[0] == item)
        else:
            qty_diff = qty
        posting_date = str(entry.posting_date)
        if posting_date < FROM_DATE or (posting_date == FROM_DATE and entry.reconciliation_purpose == "Opening Stock"):
            bucket = stock.setdefault(item, blank_stock(entry))
            bucket["opening_qty"] += qty_diff
            bucket["opening_value"] += value
        elif FROM_DATE <= str(entry.posting_date) <= TO_DATE:
            bucket = stock.setdefault(item, blank_stock(entry))
            if qty_diff >= 0:
                bucket["in_qty"] += qty_diff
                bucket["in_value"] += value
            else:
                bucket["out_qty"] += abs(qty_diff)
                bucket["out_value"] += abs(value)
        state["qty"] += qty_diff
        state["value"] += value
        latest[key] = (state["qty"], state["value"], entry.valuation_rate or 0)
        if entry.voucher_type == "Stock Entry" and posting_date >= FROM_DATE and posting_date <= TO_DATE and qty > 0:
            if entry.stock_entry_purpose == "Manufacture":
                row = manufacturing.setdefault(item, {"item_name": entry.item_name, "qty": 0, "value": 0})
                row["qty"] += qty
                row["value"] += value
    for item in stock:
        rows = [v for k, v in latest.items() if k[0] == item]
        stock[item]["balance_qty"] = sum(v[0] for v in rows)
        stock[item]["balance_value"] = sum(v[1] for v in rows)
        stock[item]["valuation_rate"] = rows[-1][2] if rows else 0
    return stock, manufacturing


def blank_stock(entry):
    return {"item_name": entry.item_name, "item_group": entry.item_group, "warehouse": entry.warehouse, "account": entry.account_name or "", "acc_code": entry.account_code or "", "opening_qty": 0, "opening_value": 0, "in_qty": 0, "in_value": 0, "out_qty": 0, "out_value": 0, "balance_qty": 0, "balance_value": 0, "valuation_rate": 0}


def get_gl_balances():
    return {row.account: row.balance for row in frappe.db.sql("SELECT account, SUM(debit - credit) AS balance FROM `tabGL Entry` WHERE company = %s AND posting_date <= %s AND is_cancelled = 0 GROUP BY account", (COMPANY, TO_DATE), as_dict=True)}


def append_sheet(sheet, headers, rows):
    sheet.append(headers)
    for item_code, data in sorted(rows.items()):
        if headers == STOCK_HEADERS:
            sheet.append([item_code, data["item_name"], data["item_group"], data["warehouse"], data["account"], data["acc_code"], data["balance_qty"], data["balance_value"], data["opening_qty"], data["opening_value"], data["in_qty"], data["in_value"], data["out_qty"], data["out_value"], data["valuation_rate"]])
        else:
            row = sheet.max_row + 1
            sheet.append([item_code, data["item_name"], data["qty"], data["value"], None])
            sheet.cell(row, 5, f"=IFERROR(D{row}/C{row},0)")


if __name__ == "__main__":
    execute()
