import openpyxl
import frappe
from frappe.utils import flt
from erpnext.stock.doctype.stock_reconciliation.stock_reconciliation import get_items

EXCEL_PATH = "/home/frappe/STOCK_BALANCE_REVALUATION_V2.xlsx"


COMPANY = "Greenphyto Pte Ltd"
WAREHOUSE = "Finished Goods - GPL"
ITEM_GROUP = "Products"
FROM_DATE = "2026-01-01"
TO_DATE = "2026-08-15"


def execute():
    rates, item_codes = get_excel_rates_and_items()
    batches = []
    for item_code in item_codes:
        batches.extend(get_items(WAREHOUSE, TO_DATE, "23:59:00", COMPANY, item_code, True))
    expense_account = frappe.db.get_value("Company", COMPANY, "stock_adjustment_account")
    cost_center = frappe.db.get_value("Company", COMPANY, "cost_center")

    reco = frappe.new_doc("Stock Reconciliation")
    reco.purpose = "Stock Reconciliation"
    reco.posting_date = TO_DATE
    reco.posting_time = "23:59:00"
    reco.set_posting_time = 1
    reco.company = COMPANY
    reco.set_warehouse = WAREHOUSE
    reco.expense_account = expense_account
    reco.cost_center = cost_center

    for batch in batches:
        new_rate = rates.get(batch.get("item_code"))
        if (
            new_rate is None
            or flt(batch.get("qty"), 2) <= 0
            or flt(batch.get("qty"), 2) * flt(new_rate, 2) <= 0
            or flt(batch.get("valuation_rate"), 2) == flt(new_rate, 2)
        ):
            continue
        reco.append("items", {
            "item_code": batch.get("item_code"),
            "warehouse": WAREHOUSE,
            "batch_no": batch.get("batch_no"),
            "qty": flt(batch.get("qty"), 2),
            "valuation_rate": flt(new_rate, 2),
        })

    if not reco.items:
        frappe.msgprint("No active batch rate changes found")
        return

    reco.insert()
    frappe.db.commit()
    frappe.msgprint(f"Stock Reconciliation draft: {reco.name}\nItems: {len(reco.items)}")


def get_excel_rates_and_items():
    workbook = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    stock_sheet = workbook["Stock Balance"]
    manufacturing_sheet = workbook["Manufacturing Value"]

    eligible_items = {
        row[0]: flt(row[6], 2)
        for row in stock_sheet.iter_rows(min_row=2, values_only=True)
        if row[0] and flt(row[6], 2) > 0
    }
    rates = {}
    for row in manufacturing_sheet.iter_rows(min_row=2, values_only=True):
        item_code, _, qty, value = row[:4]
        if item_code in eligible_items and flt(qty, 2) > 0:
            rates[item_code] = flt(value / qty, 2)

    return rates, set(rates)


def get_manufacturing_rates():
    rows = frappe.db.sql("""
        SELECT sle.item_code,
               SUM(sle.actual_qty) AS qty,
               SUM(sle.stock_value_difference) AS value
        FROM `tabStock Ledger Entry` sle
        INNER JOIN `tabItem` i ON i.name = sle.item_code
        INNER JOIN `tabStock Entry` se ON se.name = sle.voucher_no
        WHERE sle.company = %s
          AND sle.warehouse = %s
          AND i.item_group = %s
          AND sle.posting_date >= %s
          AND sle.posting_date <= %s
          AND sle.voucher_type = 'Stock Entry'
          AND se.purpose = 'Manufacture'
          AND sle.actual_qty > 0
          AND sle.is_cancelled = 0
        GROUP BY sle.item_code
    """, (COMPANY, WAREHOUSE, ITEM_GROUP, FROM_DATE, TO_DATE), as_dict=True)
    return {
        row.item_code: flt(row.value / row.qty, 2)
        for row in rows
        if flt(row.qty)
    }


def get_active_batches(item_codes):
    if not item_codes:
        return []
    placeholders = ", ".join(["%s"] * len(item_codes))
    rows = frappe.db.sql(f"""
        SELECT item_code, batch_no, qty_after_transaction AS qty, valuation_rate
        FROM (
            SELECT sle.item_code, sle.batch_no, sle.qty_after_transaction,
                   sle.valuation_rate,
                   ROW_NUMBER() OVER (
                       PARTITION BY sle.item_code, sle.batch_no
                       ORDER BY sle.posting_date DESC, sle.posting_time DESC,
                                sle.creation DESC, sle.name DESC
                   ) AS row_number
            FROM `tabStock Ledger Entry` sle
            INNER JOIN `tabItem` i ON i.name = sle.item_code
            WHERE sle.company = %s
              AND sle.warehouse = %s
              AND i.item_group = %s
              AND sle.item_code IN ({placeholders})
              AND sle.posting_date <= %s
              AND sle.is_cancelled = 0
              AND sle.batch_no IS NOT NULL
              AND sle.batch_no != ''
        ) latest
        WHERE row_number = 1 AND qty_after_transaction > 0
        ORDER BY item_code, batch_no
    """, (COMPANY, WAREHOUSE, ITEM_GROUP, *item_codes, TO_DATE), as_dict=True)
    return rows
